#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Content Calendar Template Generator — Đa ngành, liên kết động Calendar <-> Pillar
===================================================================================

MỤC TIÊU THIẾT KẾ (đáp ứng yêu cầu tái sử dụng cho nhiều ngành hàng):

1. Tách CẤU HÌNH khỏi LOGIC:
   - Sheet "Config" chứa toàn bộ thông tin đặc thù ngành/thương hiệu (năm, tháng,
     tên brand, persona, danh sách Pillar, tên kênh, màu sắc...).
   - Muốn dùng cho ngành khác: chỉ cần sửa biến CONFIG ở đầu file (hoặc sửa trực
     tiếp sheet Config sau khi xuất ra), KHÔNG cần sửa logic sinh sheet/công thức.

2. Sheet Pillar là NGUỒN DỮ LIỆU DUY NHẤT (single source of truth):
   - Cột "Ngày đăng" là giá trị nhập trực tiếp (date), không phải kết quả tính
     cứng một lần bằng Python rồi ghi chết vào ô.
   - Cột "Tuần" là CÔNG THỨC Excel (dựa theo ngày đầu tháng lấy từ Config) để
     luôn tự đúng nếu người dùng sửa ngày đăng trực tiếp trong Excel.

3. Sheet Content Calendar KHÔNG chứa giá trị tĩnh — mỗi ô ngày là CÔNG THỨC
   TEXTJOIN + IF (mảng, tương thích Excel 2016/2019, không dùng FILTER/XLOOKUP
   chỉ có ở 365) quét toàn bộ các sheet Pillar theo đúng ngày. Nếu người dùng
   thêm/sửa/xóa dòng trong sheet Pillar, Calendar tự tính lại — không cần chạy
   lại script Python.

4. Tuần buffer: nếu số tuần lịch trong tháng > số Content Pillar, tuần dư ra
   KHÔNG bị gán lại cho Pillar 1 (gây trải Pillar 1 ra 2 tuần cách xa nhau).
   Thay vào đó, tuần dư trở thành "Tuần Buffer / Tổng hợp đa Pillar" — vừa
   dùng để chèn buffer newsjacking, vừa hiển thị placeholder gợi ý.

5. Buffer newsjacking 10-15%: khi phân bổ ngày cho từng Pillar, luôn chừa lại
   một tỷ lệ ngày trống trong tuần (không gán cluster) làm buffer thời sự,
   đúng theo nguyên tắc "Content Calendar 3H" (mục 2.5 tài liệu Module 2).

6. Dashboard đối chiếu tỷ lệ 3H thực tế với baseline bằng công thức COUNTIF
   tham chiếu trực tiếp sheet Pillar (không đếm cứng bằng Python).

Đây là TEMPLATE TRỐNG (chưa có dữ liệu cluster mẫu) — điền dữ liệu thật vào
PILLAR_DATA_RAW hoặc trực tiếp trong Excel để sử dụng cho thương hiệu/ngành hàng cụ thể.
Nguồn tham chiếu chiến lược: Module 2 - Chiến lược Nội dung (calendar-3H.md).
"""

import datetime
import calendar
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.comments import Comment


def cf_fill(hex_color):
    """
    Tạo PatternFill dùng cho Conditional Formatting (dxf).

    QUAN TRỌNG: PatternFill(fgColor="RRGGBB") (chỉ 6 ký tự hex, không alpha) khiến
    openpyxl tự thêm tiền tố alpha "00" (hoàn toàn trong suốt) khi ghi vào dxf —
    trong ngữ cảnh Conditional Formatting, Excel THẬT sự tôn trọng alpha này nên
    fill trở nên VÔ HÌNH (không giống PatternFill áp trực tiếp lên cell thường,
    nơi Excel bỏ qua alpha và luôn hiển thị màu đúng). Luôn dùng "FF" + hex 6 ký
    tự (alpha đầy đủ/opaque) khi tạo fill cho CellIsRule/FormulaRule để tránh lỗi
    "màu không hiển thị dù rule đã áp dụng đúng ô".
    """
    argb = f"FF{hex_color}" if len(hex_color) == 6 else hex_color
    return PatternFill(fill_type="solid", fgColor=argb, bgColor=argb)

# ============================================================================
# 0. CONFIG — TOÀN BỘ THÔNG TIN ĐẶC THÙ NGÀNH/THƯƠNG HIỆU NẰM Ở ĐÂY
#    Muốn tái sử dụng cho ngành khác: chỉ sửa khối CONFIG này (và dữ liệu
#    cluster ở mục 1), không cần sửa logic phía dưới.
# ============================================================================

OUT_PATH = "/Users/phamvannam/Documents/GitHub/hello-affiliate/content-marketing/code/Content_Calendar_Template.xlsx"

CONFIG = {
    "nganh": "[Điền tên ngành hàng]",
    "brand": "[Điền tên thương hiệu]",
    "persona": "[Điền tên/mô tả Persona mục tiêu]",
    "year": 2026,
    "month": 9,
    "buffer_pct": 0.12,  # 12% ngày mỗi Pillar chừa trống cho newsjacking (khung 10-15%)
    "h3_baseline": {  # baseline tỷ lệ 3H đề xuất (mục 2.4 tài liệu) để Dashboard đối chiếu
        "Hero": (0.05, 0.10),
        "Hub": (0.40, 0.50),
        "Hygiene": (0.40, 0.55),
    },
    "channels": ["Facebook", "Zalo OA", "Website", "TikTok", "Email"],
    "funnel": ["TOFU", "MOFU", "BOFU"],
    "status_list": ["Draft", "Review", "Approved", "Published"],
    "h3_list": ["Hero", "Hub", "Hygiene"],
}

# Danh sách Pillar: đổi/thêm/xóa Pillar ở đây khi áp dụng cho thương hiệu/ngành cụ thể.
# key = tên sheet (ngắn, không dấu, để dùng trong công thức Excel không lỗi),
# title = tiêu đề đầy đủ hiển thị, color = màu header đậm, light = màu nền nhạt.
PILLARS = [
    dict(key="Pillar 1", title="PILLAR 1: [Điền tên chủ đề Pillar 1]",
         color="548235", light="C6E0B4"),
    dict(key="Pillar 2", title="PILLAR 2: [Điền tên chủ đề Pillar 2]",
         color="BF8F00", light="FFE699"),
    dict(key="Pillar 3", title="PILLAR 3: [Điền tên chủ đề Pillar 3]",
         color="2E75B6", light="BDD7EE"),
    dict(key="Pillar 4", title="PILLAR 4: [Điền tên chủ đề Pillar 4]",
         color="C55A11", light="F8CBAD"),
]
PILLAR_SHEETS = [p["key"] for p in PILLARS]
PILLAR_TITLES = {p["key"]: p["title"] for p in PILLARS}
PILLAR_HEADER_COLORS = {p["key"]: p["color"] for p in PILLARS}
PILLAR_COLORS = {p["key"]: p["light"] for p in PILLARS}

BUFFER_SHEET_KEY = "Tuan Buffer"
BUFFER_TITLE = "TUẦN BUFFER: Tổng hợp đa Pillar / Newsjacking"
BUFFER_COLOR = "7F7F7F"
BUFFER_LIGHT = "D9D9D9"

H3_COLORS = {"Hero": "FF0000", "Hub": "2E75B6", "Hygiene": "70AD47"}
H3_FILL = {"Hero": "FFC7CE", "Hub": "BDD7EE", "Hygiene": "E2EFDA"}

YEAR = CONFIG["year"]
MONTH = CONFIG["month"]
PERSONA = CONFIG["persona"]
BRAND = CONFIG["brand"]
CHANNELS = CONFIG["channels"]
FUNNEL = CONFIG["funnel"]
STATUS_LIST = CONFIG["status_list"]
H3_LIST = CONFIG["h3_list"]
BUFFER_PCT = CONFIG["buffer_pct"]

# ============================================================================
# 1. DỮ LIỆU CONTENT CLUSTER (nội dung thô — chưa gán ngày cụ thể).
#    Đây là phần "dữ liệu ngành" — khi áp dụng cho ngành khác, thay khối này.
# ============================================================================

# ============================================================================
# 1. DỮ LIỆU CONTENT CLUSTER — FILE NÀY LÀ TEMPLATE TRỐNG (không có nội dung mẫu).
#    Mỗi Pillar hiện là danh sách RỖNG. Để dùng thật, điền các dict cluster vào
#    đúng khóa Pillar tương ứng theo mẫu cấu trúc dưới đây (xem sheet "Huong Dan"
#    trong file Excel xuất ra để biết chi tiết từng trường ý nghĩa gì), rồi chạy
#    lại script — hoặc điền trực tiếp vào sheet Pillar trong Excel (khuyến nghị,
#    vì Ngày đăng/Tuần/Deadline/Calendar đều tự tính bằng công thức, không cần
#    chạy lại script để cập nhật nội dung).
#
#    Mẫu cấu trúc 1 cluster (tham khảo, KHÔNG bắt buộc dùng script để nhập):
#    dict(ten="...", loai="Hero|Hub|Hygiene", funnel="TOFU|MOFU|BOFU",
#         tukhoa="...", tone="...", do_dai="...", outline="...", cta="...",
#         kenh="Facebook|Zalo OA|Website|TikTok|Email", status="Draft|Review|Approved|Published")
# ============================================================================

PILLAR_DATA_RAW = {
    "Pillar 1": [],
    "Pillar 2": [],
    "Pillar 3": [],
    "Pillar 4": [],
}

# ============================================================================
# 2. TÍNH LỊCH THEO TUẦN + GÁN PILLAR/BUFFER CHO TỪNG TUẦN
# ============================================================================

first_weekday, num_days = calendar.monthrange(YEAR, MONTH)  # 0=Mon..6=Sun

weeks = []  # list các list ngày-trong-tháng, mỗi phần tử là 1 tuần lịch (Mon->Sun)
current_week = []
for day in range(1, num_days + 1):
    wd = (first_weekday + (day - 1)) % 7
    current_week.append(day)
    if wd == 6 or day == num_days:
        weeks.append(current_week)
        current_week = []

num_weeks = len(weeks)
num_pillars = len(PILLAR_SHEETS)

print(f"Tháng {MONTH}/{YEAR} có {num_days} ngày, chia thành {num_weeks} tuần lịch, {num_pillars} Pillar.")
for i, w in enumerate(weeks, start=1):
    print(f"  Tuần {i}: ngày {w[0]}-{w[-1]} ({len(w)} ngày)")

# Gán Pillar cho từng tuần THEO THỨ TỰ, không lặp vòng (tránh 1 Pillar bị trải
# ra 2 tuần cách xa nhau trong tháng). Nếu số tuần > số Pillar, các tuần dư ở
# CUỐI tháng trở thành "Tuần Buffer" (đa Pillar / newsjacking) — lựa chọn (A).
week_to_pillar = {}
for i in range(num_weeks):
    if i < num_pillars:
        week_to_pillar[i + 1] = PILLAR_SHEETS[i]
    else:
        week_to_pillar[i + 1] = BUFFER_SHEET_KEY

buffer_weeks = [w for w, p in week_to_pillar.items() if p == BUFFER_SHEET_KEY]
if buffer_weeks:
    print(f"  -> Tuần buffer (đa Pillar/newsjacking): {buffer_weeks}")

pillar_to_weeks = defaultdict(list)
for wk_num, pillar in week_to_pillar.items():
    if pillar != BUFFER_SHEET_KEY:
        pillar_to_weeks[pillar].append(wk_num)



# ---- Phân bổ ngày đăng cho từng Pillar, CÓ CHỪA BUFFER 10-15% NGÀY TRỐNG ----
PILLAR_DATA = {}
cluster_registry = []
buffer_info = {}

RESP_CYCLE = ["Writer A", "Writer B", "Designer C", "Writer A", "Content Lead", "Writer B", "Designer C", "Writer A"]
KPI_BY_3H = {
    "Hero": "Reach > 50,000; Brand Lift +10%; Share of Voice",
    "Hub": "Engagement rate > 5%; Returning visitor +; Watch time",
    "Hygiene": "Organic traffic; Time on page > 1'30\"; Giảm câu hỏi CSKH",
}

def pick_evenly_spaced(day_list, k):
    """Chọn k phần tử trải ĐỀU trên day_list (đã sort) bằng bước nhảy cố định,
    dùng để chọn vị trí buffer xen kẽ trong tuần (không dồn về đầu/cuối)."""
    day_list = sorted(day_list)
    m = len(day_list)
    if k <= 0 or m == 0:
        return []
    if k >= m:
        return list(day_list)
    step = m / k
    chosen = sorted({day_list[min(int((i + 0.5) * step), m - 1)] for i in range(k)})
    if len(chosen) < k:
        remaining = [d for d in day_list if d not in chosen]
        for d in remaining:
            chosen.append(d)
            if len(chosen) == k:
                break
        chosen = sorted(chosen)
    return chosen


# ---- Phân bổ ngày đăng cho từng Pillar, CHỌN VỊ TRÍ BUFFER XEN KẼ TRƯỚC ----
PILLAR_DATA = {}
cluster_registry = []
buffer_info = {}

RESP_CYCLE = ["Writer A", "Writer B", "Designer C", "Writer A", "Content Lead", "Writer B", "Designer C", "Writer A"]
KPI_BY_3H = {
    "Hero": "Reach > 50,000; Brand Lift +10%; Share of Voice",
    "Hub": "Engagement rate > 5%; Returning visitor +; Watch time",
    "Hygiene": "Organic traffic; Time on page > 1'30\"; Giảm câu hỏi CSKH",
}

for pillar in PILLAR_SHEETS:
    items = PILLAR_DATA_RAW[pillar]
    assigned_weeks = pillar_to_weeks[pillar]

    all_days = []
    for wk in assigned_weeks:
        all_days.extend(weeks[wk - 1])
    all_days_sorted = sorted(all_days)

    n = len(items)
    total_slots = len(all_days_sorted)
    # BUFFER LÀ RÀNG BUỘC BẮT BUỘC (không phải phần dư sau khi xếp hết bài) VÀ PHẢI
    # XEN KẼ ĐỀU trong tuần (không dồn về cuối tuần) — để mỗi buffer có thể dùng phản
    # ứng kịp thời sự/trending bất kể xảy ra vào đầu, giữa hay cuối tuần.
    min_buffer_days = max(1, round(total_slots * BUFFER_PCT)) if total_slots >= 4 else 0

    # Bước 1: chọn TRƯỚC vị trí các ngày buffer, trải đều trên toàn bộ all_days_sorted
    # (không ưu tiên/né riêng Chủ Nhật) để buffer xen kẽ xuất hiện rải rác trong tuần.
    buffer_days_of_pillar = pick_evenly_spaced(all_days_sorted, min_buffer_days)

    # Bước 2: các ngày còn lại (sau khi trừ buffer) là "pool" để xếp cluster.
    pool = sorted(set(all_days_sorted) - set(buffer_days_of_pillar))
    usable_slots = len(pool)

    if n > usable_slots:
        print(f"  !! CẢNH BÁO [{pillar}]: có {n} cluster nhưng chỉ {usable_slots} ngày khả dụng sau khi "
              f"chừa {min_buffer_days} ngày buffer xen kẽ (tổng {total_slots} ngày trong tuần được gán). "
              f"Một số ngày sẽ phải đăng nhiều hơn 1 cluster để vẫn giữ đúng buffer — nên xem xét giảm số "
              f"cluster/Pillar hoặc gán thêm tuần cho Pillar này.")

    if pool and len(pool) >= n:
        # Trải đều n bài trên "pool" ngày khả dụng (đã trừ buffer) bằng bước nhảy đều.
        chosen_days = pick_evenly_spaced(pool, n)
    elif pool:
        # Nhiều bài hơn số ngày khả dụng (đã trừ buffer) -> lặp vòng qua pool
        # (nhiều bài/ngày) — buffer vẫn được giữ nguyên, không bị lấn.
        chosen_days = [pool[i % len(pool)] for i in range(n)]
    else:
        # Không còn ngày khả dụng nào sau khi trừ buffer (tuần quá ngắn) -> đành
        # phải dùng toàn bộ ngày trong tuần (buffer = 0 cho trường hợp đặc biệt này).
        chosen_days = [all_days_sorted[i % len(all_days_sorted)] for i in range(n)] if all_days_sorted else []
        buffer_days_of_pillar = []

    new_items = []
    for idx, item in enumerate(items):
        day_num = chosen_days[idx] if idx < len(chosen_days) else (chosen_days[-1] if chosen_days else 1)
        ngay = datetime.date(YEAR, MONTH, day_num)
        new_item = dict(item)
        new_item["ngay"] = ngay
        new_items.append(new_item)
    PILLAR_DATA[pillar] = new_items
    buffer_info[pillar] = buffer_days_of_pillar

total_buffer_days = sum(len(v) for v in buffer_info.values())
print(f"  -> Tổng số ngày buffer newsjacking (trong các tuần Pillar): {total_buffer_days} ngày "
      f"(mục tiêu ~{BUFFER_PCT:.0%} mỗi Pillar).")

# ============================================================================
# 3. STYLE HELPERS
# ============================================================================

THIN = Side(style="thin", color="B7B7B7")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)

WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header_row(ws, row, ncols, fill_hex):
    fill = PatternFill("solid", fgColor=fill_hex)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER_ALL


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# 4. TẠO WORKBOOK
# ============================================================================

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Số dòng dữ liệu tối đa dự trù trên mỗi sheet Pillar (dùng cho range công thức
# cố định — người dùng có thể thêm dòng cluster trong phạm vi này mà không cần
# sửa công thức ở sheet Calendar/Dashboard).
MAX_ROWS_PER_PILLAR = 100
HEADER_ROW_PILLAR = 4  # header nằm ở dòng 4 (sau Title, Subtitle, dòng trống)
FIRST_DATA_ROW = HEADER_ROW_PILLAR + 1
LAST_DATA_ROW = HEADER_ROW_PILLAR + MAX_ROWS_PER_PILLAR

PILLAR_COLUMNS = [
    ("Mã Cluster", 12),          # A
    ("Tên bài/Chủ đề", 34),      # B
    ("Loại 3H", 10),             # C
    ("Mục tiêu phễu\n(TOFU/MOFU/BOFU)", 14),  # D
    ("Persona mục tiêu", 26),    # E
    ("Từ khóa chính", 26),       # F
    ("Tone giọng điệu", 26),     # G
    ("Độ dài mong muốn", 20),    # H
    ("Outline/Dàn ý chính", 40), # I
    ("CTA mong muốn", 26),       # J
    ("Kênh đăng", 12),           # K
    ("Ngày đăng", 12),           # L  <-- input trực tiếp, nguồn sự thật
    ("Tuần (auto)", 10),         # M  <-- công thức, tự tính từ Ngày đăng + Config
    ("Deadline", 12),            # N
    ("Người phụ trách", 16),     # O
    ("Trạng thái", 12),          # P
    ("KPI dự kiến", 22),         # Q
]
COL_NGAY = 12   # L
COL_TUAN = 13   # M
COL_LOAI = 3    # C

# ----------------------------------------------------------------------------
# 4.1 Sheet "Config" — tách cấu hình khỏi logic, để tái sử dụng cho ngành khác
# ----------------------------------------------------------------------------

ws_cfg = wb.create_sheet("Config")
ws_cfg.sheet_view.showGridLines = False
ws_cfg.column_dimensions["A"].width = 26
ws_cfg.column_dimensions["B"].width = 46
ws_cfg.column_dimensions["C"].width = 70

ws_cfg.merge_cells("A1:C1")
c = ws_cfg.cell(row=1, column=1, value="CONFIG — CẤU HÌNH THƯƠNG HIỆU / NGÀNH HÀNG (sửa ở đây để tái sử dụng template)")
c.font = Font(size=13, bold=True, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="203864")
c.alignment = Alignment(horizontal="left", vertical="center")
ws_cfg.row_dimensions[1].height = 24

cfg_rows = [
    ("nganh", "Ngành hàng", CONFIG["nganh"], "Chỉ mang tính mô tả, không ảnh hưởng công thức."),
    ("brand", "Thương hiệu (Brand)", BRAND, "Hiển thị tại tiêu đề Calendar/Dashboard."),
    ("persona", "Persona mục tiêu", PERSONA, "Điền persona chính; nếu có nhiều persona, tạo thêm dòng trong Pillar."),
    ("year", "Năm kế hoạch", YEAR, "Số nguyên, ví dụ 2026. Đổi giá trị này + Tháng sẽ tự đổi lưới Calendar khi chạy lại script."),
    ("month", "Tháng kế hoạch", MONTH, "Số 1-12."),
    ("first_day", "Ngày đầu tháng", datetime.date(YEAR, MONTH, 1), "Dùng làm mốc cho công thức tính Tuần (auto) ở các sheet Pillar."),
    ("buffer_pct", "Buffer newsjacking (%)", BUFFER_PCT, "Tỷ lệ ngày/tuần chừa trống không gán cluster, dùng bắt trend (khung đề xuất 10-15%)."),
]
CFG_ROW = {}  # key -> số dòng thực tế trong sheet Config, dùng để build công thức tham chiếu đúng
r = 3
ws_cfg.cell(row=r, column=1, value="THÔNG TIN CHUNG").font = Font(bold=True, size=11, color="203864")
r += 1
for key, label, value, note in cfg_rows:
    CFG_ROW[key] = r
    ws_cfg.cell(row=r, column=1, value=label).font = BOLD_FONT
    vcell = ws_cfg.cell(row=r, column=2, value=value)
    if isinstance(value, datetime.date):
        vcell.number_format = "dd/mm/yyyy"
    if isinstance(value, float):
        vcell.number_format = "0%"
    ws_cfg.cell(row=r, column=3, value=note).font = Font(italic=True, size=9, color="595959")
    for cc in (1, 2, 3):
        ws_cfg.cell(row=r, column=cc).border = BORDER_ALL
        ws_cfg.cell(row=r, column=cc).alignment = WRAP_TOP
    r += 1

r += 1
ws_cfg.cell(row=r, column=1, value="DANH SÁCH CONTENT PILLAR").font = Font(bold=True, size=11, color="203864")
r += 1
for c_idx, h in enumerate(["Tên sheet Pillar", "Tiêu đề đầy đủ", "Màu header (hex)"], start=1):
    cell = ws_cfg.cell(row=r, column=c_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.border = BORDER_ALL
    cell.alignment = CENTER
r += 1
for p in PILLARS:
    ws_cfg.cell(row=r, column=1, value=p["key"]).border = BORDER_ALL
    ws_cfg.cell(row=r, column=2, value=p["title"]).border = BORDER_ALL
    cell = ws_cfg.cell(row=r, column=3, value=p["color"])
    cell.border = BORDER_ALL
    cell.fill = PatternFill("solid", fgColor=p["color"])
    cell.font = Font(color="FFFFFF", bold=True)
    r += 1
ws_cfg.cell(row=r, column=1, value=BUFFER_SHEET_KEY).border = BORDER_ALL
ws_cfg.cell(row=r, column=2, value=BUFFER_TITLE).border = BORDER_ALL
cell = ws_cfg.cell(row=r, column=3, value=BUFFER_COLOR)
cell.border = BORDER_ALL
cell.fill = PatternFill("solid", fgColor=BUFFER_COLOR)
cell.font = Font(color="FFFFFF", bold=True)
r += 2

ws_cfg.cell(row=r, column=1, value="DANH MỤC DÙNG CHO DROPDOWN (sheet Pillar)").font = Font(bold=True, size=11, color="203864")
r += 1
for label, values in [("Kênh (Channels)", CHANNELS), ("Mục tiêu phễu (Funnel)", FUNNEL),
                       ("Trạng thái (Status)", STATUS_LIST), ("Loại 3H", H3_LIST)]:
    ws_cfg.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws_cfg.cell(row=r, column=2, value=", ".join(values)).border = BORDER_ALL
    ws_cfg.cell(row=r, column=1).border = BORDER_ALL
    r += 1

r += 1
note_text = ("Ghi chú: Đây là bảng tham chiếu hiển thị. Do giới hạn của Excel công thức tĩnh, "
             "danh sách dropdown thực tế trong sheet Pillar được thiết lập sẵn theo đúng danh mục này "
             "khi sinh file. Muốn đổi danh mục cho ngành khác, sửa CONFIG trong script rồi chạy lại.")
ws_cfg.cell(row=r, column=1, value=note_text)
ws_cfg.cell(row=r, column=1).font = Font(italic=True, size=9, color="C00000")
ws_cfg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws_cfg.cell(row=r, column=1).alignment = WRAP_TOP
ws_cfg.row_dimensions[r].height = 32

# ----------------------------------------------------------------------------
# 4.2 Tạo sheet Pillar (Ngày đăng = input, Tuần = công thức)
# ----------------------------------------------------------------------------

all_pillar_sheet_keys = PILLAR_SHEETS + [BUFFER_SHEET_KEY]

for sheet_name in all_pillar_sheet_keys:
    ws = wb.create_sheet(sheet_name)
    is_buffer = sheet_name == BUFFER_SHEET_KEY
    header_color = BUFFER_COLOR if is_buffer else PILLAR_HEADER_COLORS[sheet_name]
    light_color = BUFFER_LIGHT if is_buffer else PILLAR_COLORS[sheet_name]
    title_text = BUFFER_TITLE if is_buffer else PILLAR_TITLES[sheet_name]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PILLAR_COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = TITLE_FONT
    title_cell.fill = PatternFill("solid", fgColor=header_color)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if is_buffer:
        weeks_label = ", ".join(f"Tuần {w}" for w in buffer_weeks) if buffer_weeks else "(không có tuần buffer trong tháng này)"
        sub_text = (f"Dùng để chèn nội dung newsjacking/thời sự phát sinh, hoặc tổng hợp bài xen ngang từ nhiều Pillar. "
                    f"Tuần được xếp làm buffer: {weeks_label}")
    else:
        assigned_weeks = pillar_to_weeks[sheet_name]
        weeks_label = ", ".join(f"Tuần {w}" for w in assigned_weeks)
        sub_text = f"Persona: {PERSONA}  |  Thương hiệu: {BRAND}  |  Pillar này đăng vào: {weeks_label}  (Xem đầy đủ cấu hình tại sheet Config)"

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(PILLAR_COLUMNS))
    sub_cell = ws.cell(row=2, column=1, value=sub_text)
    sub_cell.font = Font(italic=True, size=10, color="404040")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    header_row = HEADER_ROW_PILLAR
    for c_idx, (name, width) in enumerate(PILLAR_COLUMNS, start=1):
        ws.cell(row=header_row, column=c_idx, value=name)
    style_header_row(ws, header_row, len(PILLAR_COLUMNS), header_color)
    ws.row_dimensions[header_row].height = 32

    autofit(ws, [w for _, w in PILLAR_COLUMNS])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3).coordinate

    items = [] if is_buffer else PILLAR_DATA[sheet_name]

    for idx in range(MAX_ROWS_PER_PILLAR):
        r_ = header_row + 1 + idx
        has_data = idx < len(items)
        item = items[idx] if has_data else None

        # --- Cột L: Ngày đăng (INPUT trực tiếp, nguồn sự thật) ---
        ngay_cell = ws.cell(row=r_, column=COL_NGAY)
        if has_data:
            ngay_cell.value = item["ngay"]
        ngay_cell.number_format = "dd/mm/yyyy"

        # --- Cột M: Tuần (auto) = công thức dựa trên Ngày đăng & ngày đầu tháng ở Config ---
        # Số tuần lịch trong tháng kể từ Thứ 2 đầu tiên chứa/trước ngày 1 (tương thích
        # Excel cũ: chỉ dùng INT/WEEKDAY, không dùng hàm 365-only).
        # Số tuần lịch (Mon->Sun) trong tháng chứa Ngày đăng, tính từ ngày đầu tháng ở
        # Config. Công thức: INT((NgayDang - NgayDauThang + WEEKDAY(NgayDauThang,2) - 1)/7) + 1
        # — khớp đúng với cách chia tuần Mon->Sun dùng trong Python (mục 2). Tương thích
        # Excel cũ: chỉ dùng INT/WEEKDAY, không dùng hàm 365-only.
        tuan_formula = (
            f'=IF(L{r_}="","",'
            f'INT((L{r_}-Config!$B${CFG_ROW["first_day"]}+WEEKDAY(Config!$B${CFG_ROW["first_day"]},2)-1)/7)+1)'
        )
        ws.cell(row=r_, column=COL_TUAN, value=tuan_formula)

        # --- Cột N: Deadline = Ngày đăng - 2 (công thức, tự theo Ngày đăng) ---
        deadline_cell = ws.cell(row=r_, column=14, value=f'=IF(L{r_}="","",L{r_}-2)')
        deadline_cell.number_format = "dd/mm/yyyy"

        row_values = {}
        if has_data:
            ma_cluster_prefix = "BUF" if is_buffer else sheet_name.split(" ")[1]
            row_values = {
                1: f"{ma_cluster_prefix}-{idx+1:02d}",
                2: item["ten"], 3: item["loai"], 4: item["funnel"],
                5: PERSONA, 6: item["tukhoa"], 7: item["tone"], 8: item["do_dai"],
                9: item["outline"], 10: item["cta"], 11: item["kenh"],
                15: RESP_CYCLE[idx % len(RESP_CYCLE)], 16: item["status"],
                17: KPI_BY_3H[item["loai"]],
            }
        for c_idx in range(1, len(PILLAR_COLUMNS) + 1):
            if c_idx in (COL_NGAY, COL_TUAN, 14):
                cell = ws.cell(row=r_, column=c_idx)
            else:
                cell = ws.cell(row=r_, column=c_idx, value=row_values.get(c_idx))
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER_ALL

        row_fill = PatternFill("solid", fgColor=light_color if has_data else "FFFFFF")
        for c_idx in range(1, len(PILLAR_COLUMNS) + 1):
            if c_idx != COL_LOAI:
                ws.cell(row=r_, column=c_idx).fill = row_fill

        if has_data:
            h3_cell = ws.cell(row=r_, column=COL_LOAI)
            h3_cell.fill = PatternFill("solid", fgColor=H3_FILL[item["loai"]])
            h3_cell.font = Font(bold=True, color=H3_COLORS[item["loai"]], size=10)
            h3_cell.alignment = CENTER

            cluster_registry.append({
                "pillar_sheet": sheet_name,
                "pillar_title": title_text,
                "row": r_, "ma_cluster": row_values[1], "ten": item["ten"], "loai": item["loai"],
                "kenh": item["kenh"], "ngay": item["ngay"], "funnel": item["funnel"], "status": item["status"],
            })
        else:
            ws.cell(row=r_, column=COL_LOAI).fill = PatternFill("solid", fgColor="FFFFFF")

    dv_3h = DataValidation(type="list", formula1='"Hero,Hub,Hygiene"', allow_blank=True)
    dv_kenh = DataValidation(type="list", formula1=f'"{",".join(CHANNELS)}"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1=f'"{",".join(STATUS_LIST)}"', allow_blank=True)
    dv_funnel = DataValidation(type="list", formula1=f'"{",".join(FUNNEL)}"', allow_blank=True)
    # Validation cho cột Ngày đăng: chỉ chấp nhận ngày hợp lệ trong khoảng năm 2000-2100,
    # kèm prompt nhắc rõ định dạng dd/mm/yyyy — giúp giảm rủi ro Excel hiểu sai thứ tự
    # ngày/tháng khi người dùng tự gõ tay (hành vi input phụ thuộc locale hệ thống của
    # Excel, không thể ép buộc hoàn toàn từ file .xlsx, nhưng prompt + error alert giúp
    # người dùng nhận biết ngay khi gõ sai).
    dv_ngay = DataValidation(
        type="date", operator="between",
        formula1="DATE(2000,1,1)", formula2="DATE(2100,12,31)",
        allow_blank=True, showInputMessage=True, showErrorMessage=True,
    )
    dv_ngay.promptTitle = "Nhập Ngày đăng"
    dv_ngay.prompt = "Nhập theo định dạng dd/mm/yyyy (ví dụ 25/12/2026). Ô đã đặt sẵn number format dd/mm/yyyy."
    dv_ngay.errorTitle = "Ngày không hợp lệ"
    dv_ngay.error = "Vui lòng nhập một ngày hợp lệ theo định dạng dd/mm/yyyy (ví dụ 25/12/2026)."
    ws.add_data_validation(dv_3h); ws.add_data_validation(dv_kenh)
    ws.add_data_validation(dv_status); ws.add_data_validation(dv_funnel)
    ws.add_data_validation(dv_ngay)
    dv_3h.add(f"C{FIRST_DATA_ROW}:C{LAST_DATA_ROW}")
    dv_kenh.add(f"K{FIRST_DATA_ROW}:K{LAST_DATA_ROW}")
    dv_status.add(f"P{FIRST_DATA_ROW}:P{LAST_DATA_ROW}")
    dv_funnel.add(f"D{FIRST_DATA_ROW}:D{LAST_DATA_ROW}")
    dv_ngay.add(f"L{FIRST_DATA_ROW}:L{LAST_DATA_ROW}")

    status_col = "P"
    ws.conditional_formatting.add(f"{status_col}{FIRST_DATA_ROW}:{status_col}{LAST_DATA_ROW}",
        CellIsRule(operator="equal", formula=['"Published"'], fill=cf_fill("C6EFCE"), font=Font(color="006100", bold=True)))
    ws.conditional_formatting.add(f"{status_col}{FIRST_DATA_ROW}:{status_col}{LAST_DATA_ROW}",
        CellIsRule(operator="equal", formula=['"Draft"'], fill=cf_fill("FFEB9C"), font=Font(color="9C6500", bold=True)))
    ws.conditional_formatting.add(f"{status_col}{FIRST_DATA_ROW}:{status_col}{LAST_DATA_ROW}",
        CellIsRule(operator="equal", formula=['"Approved"'], fill=cf_fill("DDEBF7"), font=Font(color="2E75B6", bold=True)))
    ws.conditional_formatting.add(f"{status_col}{FIRST_DATA_ROW}:{status_col}{LAST_DATA_ROW}",
        CellIsRule(operator="equal", formula=['"Review"'], fill=cf_fill("FCE4D6"), font=Font(color="C55A11", bold=True)))

    ws.sheet_view.showGridLines = False

print(f"Đã tạo {len(all_pillar_sheet_keys)} sheet Pillar (bao gồm Tuần Buffer), "
      f"tổng {len(cluster_registry)} content cluster có dữ liệu mẫu "
      f"(mỗi sheet dự trù {MAX_ROWS_PER_PILLAR} dòng có thể tự thêm/sửa/xóa).")

# ============================================================================
# 5. SHEET CONTENT CALENDAR — CÔNG THỨC ĐỘNG (TEXTJOIN + IF), KHÔNG GHI TĨNH
#    Mỗi ô ngày quét TẤT CẢ sheet Pillar + Tuần Buffer để tìm cluster có
#    Ngày đăng khớp ngày đó. Nếu Pillar sheet bị sửa/xóa dòng, ô này tự đổi.
# ============================================================================

ws_cal = wb.create_sheet("Content Calendar", 0)
ws_cal.sheet_view.showGridLines = False

weekday_names_vn = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"]
n_cols = 1 + 7
last_col_idx = n_cols

ws_cal.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_idx)
t = ws_cal.cell(row=1, column=1,
    value=f'="CONTENT CALENDAR THÁNG "&Config!$B${CFG_ROW["month"]}&"/"&Config!$B${CFG_ROW["year"]}&" — "&Config!$B${CFG_ROW["brand"]}&" (liên kết động với sheet Pillar)"')
t.font = Font(size=15, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="375623")
t.alignment = Alignment(horizontal="center", vertical="center")
ws_cal.row_dimensions[1].height = 30

header_row_cal = 2
ws_cal.cell(row=header_row_cal, column=1, value="Tuần / Pillar")
for c_idx in range(1, n_cols + 1):
    cell = ws_cal.cell(row=header_row_cal, column=c_idx)
    if c_idx > 1:
        wd_idx = c_idx - 2
        is_weekend = wd_idx == 6
        cell.value = weekday_names_vn[wd_idx]
        cell.fill = PatternFill("solid", fgColor="F2F2F2" if not is_weekend else "FCE4D6")
        cell.font = Font(bold=True, color="000000" if not is_weekend else "C00000")
    else:
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.border = BORDER_ALL
ws_cal.row_dimensions[header_row_cal].height = 22

ws_cal.freeze_panes = ws_cal.cell(row=header_row_cal + 1, column=2).coordinate


def build_single_sheet_formula(sheet_key, date_cell_ref):
    """
    Công thức TEXTJOIN + IF (mảng) CHỈ quét 1 sheet Pillar để lấy nội dung của
    đúng ngày `date_cell_ref`. Đây là 1 'helper cell' — công thức được giữ NGẮN
    và ĐƠN GIẢN (chỉ 1 sheet/công thức) để tối đa hóa khả năng tương thích khi
    mở bằng các phiên bản Excel/trình đọc khác nhau, thay vì gộp chung nhiều
    sheet vào 1 công thức duy nhất rất dài.

    Range cố định FIRST_DATA_ROW:LAST_DATA_ROW nên nếu người dùng thêm/sửa/xóa
    dòng cluster trong phạm vi đó, kết quả tự cập nhật — không cần sửa công
    thức hay chạy lại script.

    Chỉ dùng TEXTJOIN + IF (có từ Excel 2016/2019), KHÔNG dùng FILTER/XLOOKUP
    (365-only), tương thích ngược theo yêu cầu.
    """
    sheet_ref = f"'{sheet_key}'"
    rng_ten = f"{sheet_ref}!$B${FIRST_DATA_ROW}:$B${LAST_DATA_ROW}"
    rng_loai = f"{sheet_ref}!$C${FIRST_DATA_ROW}:$C${LAST_DATA_ROW}"
    rng_ngay = f"{sheet_ref}!$L${FIRST_DATA_ROW}:$L${LAST_DATA_ROW}"
    # QUAN TRỌNG: TEXTJOIN là hàm mới hơn baseline Excel 2007 (chỉ có từ Excel 2016/2019).
    # openpyxl/OOXML yêu cầu các hàm này phải được ghi với tiền tố "_xlfn." trong XML để
    # Excel nhận diện đúng — nếu ghi thẳng "TEXTJOIN(...)" (không tiền tố), Excel THẬT sẽ
    # báo lỗi #NAME? vì không tìm thấy tên hàm hợp lệ (dù hiển thị đúng "TEXTJOIN" khi đọc
    # lại bằng openpyxl, vì openpyxl tự ẩn tiền tố khi đọc). Đây là quy tắc bắt buộc của
    # định dạng OOXML, không phải lựa chọn tùy chọn.
    formula = (
        f'=_xlfn.TEXTJOIN("",TRUE,IF({rng_ngay}={date_cell_ref},'
        f'"["&{rng_loai}&"] "&{rng_ten}&CHAR(10),""))'
    )
    return formula


# Build map ngày -> list cluster entries (chỉ dùng để style màu ô theo loại
# 3H chủ đạo của ngày và gắn hyperlink/comment gợi ý — nội dung TEXT thật sự
# hiển thị trong ô vẫn do CÔNG THỨC Excel tính, không phải Python ghi tĩnh).
day_map = defaultdict(list)
for c_ in cluster_registry:
    day_map[c_["ngay"].day].append(c_)

row_cursor = header_row_cal + 1
# Vùng helper (ẩn) bên phải lưới: mỗi ngày cần 1 cột "ngày mốc" + N cột (1/sheet
# Pillar+Buffer) chứa kết quả TEXTJOIN riêng của từng sheet + 1 cột "Loại 3H chủ đạo"
# (dùng để Conditional Formatting tô màu ô hiển thị tự động, KHÔNG dùng PatternFill
# tĩnh nữa — màu sẽ tự đổi đúng khi người dùng thêm/sửa/xóa cluster trong sheet Pillar,
# không chỉ đúng tại thời điểm chạy script). Bố trí theo block liên tiếp cho từng cột
# thứ (Thứ 2..CN): [ngày mốc][kq sheet1][kq sheet2]...[kq sheetN][loại 3H chủ đạo][địa chỉ link]
HELPER_BLOCK_WIDTH = 1 + len(all_pillar_sheet_keys) + 1 + 1  # ngày mốc + N kết quả/sheet + loại 3H + địa chỉ link
HELPER_COL_START = n_cols + 3

for wk_idx, day_list in enumerate(weeks, start=1):
    pillar = week_to_pillar[wk_idx]
    is_buffer_week = pillar == BUFFER_SHEET_KEY
    header_color = BUFFER_COLOR if is_buffer_week else PILLAR_HEADER_COLORS[pillar]
    light_color = BUFFER_LIGHT if is_buffer_week else PILLAR_COLORS[pillar]
    label_title = BUFFER_TITLE if is_buffer_week else PILLAR_TITLES[pillar]

    row_daynum = row_cursor
    row_content = row_cursor + 1

    ws_cal.merge_cells(start_row=row_daynum, start_column=1, end_row=row_content, end_column=1)
    short_label = label_title.replace("PILLAR ", "P").replace("TUẦN BUFFER: ", "BUFFER: ")
    label_cell = ws_cal.cell(row=row_daynum, column=1, value=f"TUẦN {wk_idx}\n{short_label}")
    label_cell.font = Font(bold=True, size=10, color="FFFFFF")
    label_cell.fill = PatternFill("solid", fgColor=header_color)
    label_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    label_cell.border = BORDER_ALL

    day_by_wd = {}
    for day in day_list:
        wd = (first_weekday + (day - 1)) % 7
        day_by_wd[wd] = day

    for wd in range(7):
        col = 2 + wd
        day_cell = ws_cal.cell(row=row_daynum, column=col)
        content_cell = ws_cal.cell(row=row_content, column=col)
        day_cell.border = BORDER_ALL
        content_cell.border = BORDER_ALL

        if wd in day_by_wd:
            day_num = day_by_wd[wd]
            day_cell.value = day_num
            day_cell.font = Font(bold=True, size=9, color="595959")
            day_cell.alignment = CENTER
            day_cell.fill = PatternFill("solid", fgColor=light_color)

            block_start = HELPER_COL_START + wd * HELPER_BLOCK_WIDTH
            date_helper_col = block_start
            date_helper_cell = ws_cal.cell(row=row_daynum, column=date_helper_col,
                                            value=datetime.date(YEAR, MONTH, day_num))
            date_helper_cell.number_format = "dd/mm/yyyy"
            date_helper_ref = f"${get_column_letter(date_helper_col)}${row_daynum}"

            # Mỗi sheet Pillar/Buffer có 1 cột helper riêng chứa công thức TEXTJOIN
            # đơn giản (chỉ quét 1 sheet) — giữ mỗi công thức ngắn, dễ kiểm tra lỗi.
            per_sheet_helper_refs = []
            per_sheet_loai_formulas = []
            per_sheet_row_formulas = []
            for i, sheet_key in enumerate(all_pillar_sheet_keys):
                helper_col = block_start + 1 + i
                helper_cell = ws_cal.cell(row=row_content, column=helper_col)
                sheet_formula = build_single_sheet_formula(sheet_key, date_helper_ref)
                helper_cell.value = ArrayFormula(helper_cell.coordinate, sheet_formula)
                per_sheet_helper_refs.append(f"${get_column_letter(helper_col)}${row_content}")
                # Công thức tìm loại 3H (cột C) của DÒNG ĐẦU TIÊN trong sheet này có
                # Ngày đăng khớp đúng ngày hiện tại — dùng cho helper "loại 3H chủ đạo".
                sheet_ref_q = f"'{sheet_key}'"
                rng_ngay_q = f"{sheet_ref_q}!$L${FIRST_DATA_ROW}:$L${LAST_DATA_ROW}"
                rng_loai_q = f"{sheet_ref_q}!$C${FIRST_DATA_ROW}:$C${LAST_DATA_ROW}"
                per_sheet_loai_formulas.append(
                    f'IFERROR(INDEX({rng_loai_q},MATCH({date_helper_ref},{rng_ngay_q},0)),"")'
                )
                # Công thức tính SỐ DÒNG (row) trong sheet Pillar của cluster đầu tiên khớp
                # ngày này — dùng để dựng địa chỉ HYPERLINK động (thay cho hyperlink tĩnh
                # cũ, vốn chỉ đúng tại thời điểm chạy script và không cập nhật khi người
                # dùng tự thêm/sửa/xóa cluster trực tiếp trong Excel).
                per_sheet_row_formulas.append(
                    f'IFERROR(MATCH({date_helper_ref},{rng_ngay_q},0)+{FIRST_DATA_ROW}-1,0)'
                )

            # Ô hiển thị chính: chỉ NỐI CHUỖI các helper (KHÔNG phải array formula,
            # công thức đơn giản, tương thích tối đa mọi phiên bản Excel).
            content_cell.value = "=TRIM(" + "&".join(per_sheet_helper_refs) + ")"
            content_cell.alignment = WRAP_TOP
            content_cell.font = Font(size=8, bold=True)
            content_cell.fill = PatternFill("solid", fgColor="FFFFFF")  # nền mặc định; Conditional
            # Formatting (thêm dưới) sẽ tự đè màu đúng theo loại 3H khi có cluster khớp ngày.

            # Helper "Loại 3H chủ đạo" của ngày: lấy loại 3H của cluster đầu tiên tìm
            # được (quét theo thứ tự sheet Pillar -> Buffer), dùng làm điều kiện cho
            # Conditional Formatting tô màu ô hiển thị — THAY CHO PatternFill tĩnh cũ
            # (vốn chỉ đúng tại thời điểm chạy script, không tự cập nhật khi người dùng
            # thêm/sửa/xóa cluster trực tiếp trong Excel).
            loai_helper_col = block_start + 1 + len(all_pillar_sheet_keys)
            loai_helper_cell = ws_cal.cell(row=row_content, column=loai_helper_col)
            # Lấy loại 3H của cluster ĐẦU TIÊN tìm được (theo thứ tự sheet Pillar -> Buffer)
            # bằng IF lồng tuần tự đơn giản (không dùng array formula phức tạp, tương thích
            # tối đa, dễ đọc/debug hơn cách dùng FIND/LEFT để giả lập "COALESCE").
            loai_formula = per_sheet_loai_formulas[-1]
            for f in reversed(per_sheet_loai_formulas[:-1]):
                loai_formula = f'IF({f}<>"",{f},{loai_formula})'
            loai_formula = "=" + loai_formula
            loai_helper_cell.value = loai_formula
            loai_helper_ref = f"${get_column_letter(loai_helper_col)}${row_content}"

            # Helper "Địa chỉ liên kết" của ngày: dựng chuỗi text dạng "'TênSheet'!A<dòng>"
            # cho cluster ĐẦU TIÊN tìm được (theo đúng thứ tự sheet Pillar -> Buffer, khớp
            # với thứ tự dùng ở loai_formula/content_cell) — dùng làm địa chỉ cho HYPERLINK()
            # ĐỘNG. Thay cho hyperlink TĨNH cũ (gán 1 lần lúc chạy script dựa vào dữ liệu
            # snapshot trong PILLAR_DATA_RAW) — hyperlink tĩnh không cập nhật khi người dùng
            # tự thêm/sửa/xóa cluster trực tiếp trong Excel (đây là lỗi của phiên bản trước).
            link_addr_col = block_start + 2 + len(all_pillar_sheet_keys)
            link_addr_cell = ws_cal.cell(row=row_content, column=link_addr_col)
            link_formula = f'"\'{all_pillar_sheet_keys[-1]}\'!A"&{per_sheet_row_formulas[-1]}'
            for i in range(len(all_pillar_sheet_keys) - 2, -1, -1):
                sk = all_pillar_sheet_keys[i]
                rf = per_sheet_row_formulas[i]
                link_formula = f'IF({per_sheet_loai_formulas[i]}<>"","\'{sk}\'!A"&{rf},{link_formula})'
            link_addr_cell.value = "=" + link_formula
            link_addr_ref = f"${get_column_letter(link_addr_col)}${row_content}"

            # Ô hiển thị chính giờ dùng HYPERLINK() ĐỘNG: đích đến tự tính lại theo đúng
            # sheet/dòng của cluster hiện tại — nếu người dùng thêm/sửa/xóa cluster trong
            # Pillar, link sẽ tự đổi theo, không cần chạy lại script. Nếu không có cluster
            # nào khớp ngày (loai_helper_ref rỗng), HYPERLINK trỏ về chính ô này (vô hại).
            content_cell.value = (
                f'=IF({loai_helper_ref}="",TRIM({"&".join(per_sheet_helper_refs)}),'
                f'HYPERLINK("#"&{link_addr_ref},TRIM({"&".join(per_sheet_helper_refs)})))'
            )
            content_cell.alignment = WRAP_TOP
            content_cell.font = Font(size=8, bold=True)
            content_cell.fill = PatternFill("solid", fgColor="FFFFFF")  # nền mặc định; Conditional
            # Formatting (thêm dưới) sẽ tự đè màu đúng theo loại 3H khi có cluster khớp ngày.

            content_cell.comment = Comment(
                "Nội dung, màu ô và liên kết (click để mở) đều là CÔNG THỨC tự động lấy từ sheet Pillar\n"
                "theo Ngày đăng — không phải giá trị/hyperlink tĩnh. Sửa/xóa dòng trong Pillar sẽ cập nhật ngay.\n"
                "Nếu ô đang trống: có thể là ngày BUFFER (chừa trống cho newsjacking) hoặc chưa có cluster nào cho ngày này.",
                "Content Calendar")

            # MÀU Ô TỰ ĐỘNG THEO LOẠI 3H: dùng Conditional Formatting (không dùng
            # PatternFill tĩnh) — điều kiện tham chiếu đúng ô helper "loại 3H chủ đạo"
            # (loai_helper_ref) của CHÍNH ngày này. Nhờ vậy khi người dùng thêm/sửa/xóa
            # cluster trực tiếp trong sheet Pillar, màu ô Calendar sẽ tự đổi theo ngay,
            # không chỉ đúng tại thời điểm chạy script (đây là lỗi của phiên bản trước).
            cell_addr = content_cell.coordinate
            for h3, fill_hex in H3_FILL.items():
                ws_cal.conditional_formatting.add(
                    cell_addr,
                    FormulaRule(formula=[f'{loai_helper_ref}="{h3}"'],
                                fill=cf_fill(fill_hex),
                                font=Font(size=8, bold=True, color=H3_COLORS[h3]))
                )
        else:
            day_cell.fill = PatternFill("solid", fgColor="F2F2F2")
            content_cell.fill = PatternFill("solid", fgColor="F2F2F2")

    ws_cal.row_dimensions[row_daynum].height = 16
    ws_cal.row_dimensions[row_content].height = 62
    row_cursor = row_content + 1

ws_cal.column_dimensions["A"].width = 22
for c_idx in range(2, n_cols + 1):
    ws_cal.column_dimensions[get_column_letter(c_idx)].width = 20
total_helper_cols = 7 * HELPER_BLOCK_WIDTH
for offset in range(total_helper_cols):
    helper_col = HELPER_COL_START + offset
    ws_cal.column_dimensions[get_column_letter(helper_col)].hidden = True

legend_row = row_cursor + 1
ws_cal.cell(row=legend_row, column=1, value="CHÚ GIẢI (LEGEND)").font = Font(bold=True, size=11)
legend_row += 1
legend_items = [("Hero (HE)", H3_FILL["Hero"], H3_COLORS["Hero"]),
                ("Hub (HU)", H3_FILL["Hub"], H3_COLORS["Hub"]),
                ("Hygiene (HY)", H3_FILL["Hygiene"], H3_COLORS["Hygiene"])]
for i, (label, fill_hex, font_hex) in enumerate(legend_items):
    cell = ws_cal.cell(row=legend_row + i, column=1, value=label)
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.font = Font(bold=True, color=font_hex)
    cell.border = BORDER_ALL
    ws_cal.cell(row=legend_row + i, column=2,
        value="Nội dung trong ô Calendar là CÔNG THỨC tự động — sửa/xóa dòng trong sheet Pillar sẽ cập nhật ngay, không cần chạy lại script.").font = Font(italic=True, size=9, color="595959")

legend_row2 = legend_row + len(legend_items) + 1
ws_cal.cell(row=legend_row2, column=1,
    value="Ô màu xám (nếu có) là TUẦN BUFFER — dành cho newsjacking hoặc tổng hợp nội dung xen ngang từ nhiều Pillar, không gán cố định 1 Pillar.").font = Font(italic=True, size=9, color="595959")
legend_row3 = legend_row2 + 1
pillar_week_summary = "; ".join(
    f"Tuần {w}: {(BUFFER_TITLE if p == BUFFER_SHEET_KEY else PILLAR_TITLES[p]).split(': ')[-1]}"
    for w, p in week_to_pillar.items()
)
ws_cal.cell(row=legend_row3, column=1, value="Phân bổ: " + pillar_week_summary).font = Font(italic=True, size=9, color="595959")
ws_cal.merge_cells(start_row=legend_row3, start_column=1, end_row=legend_row3, end_column=last_col_idx)

# ============================================================================
# 6. SHEET DASHBOARD — COUNTIF trực tiếp vào sheet Pillar (tự cập nhật)
#    + đánh giá ĐẠT/CHƯA ĐẠT tỷ lệ 3H so với baseline lấy từ Config.
# ============================================================================

ws_dash = wb.create_sheet("Dashboard Can Bang", 1)
ws_dash.sheet_view.showGridLines = False

ws_dash.merge_cells("A1:E1")
t2 = ws_dash.cell(row=1, column=1,
    value=f'="DASHBOARD KIỂM TRA CÂN BẰNG NỘI DUNG — THÁNG "&Config!$B${CFG_ROW["month"]}&"/"&Config!$B${CFG_ROW["year"]}')
t2.font = Font(size=14, bold=True, color="FFFFFF")
t2.fill = PatternFill("solid", fgColor="203864")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 26

ws_dash.cell(row=3, column=1, value="Tuần").font = HEADER_FONT
ws_dash.cell(row=3, column=2, value="Pillar được gán").font = HEADER_FONT
ws_dash.cell(row=3, column=3, value="Số Content Cluster").font = HEADER_FONT
ws_dash.cell(row=3, column=4, value="Đạt tối thiểu 3-5 bài?").font = HEADER_FONT
for c_idx in (1, 2, 3, 4):
    ws_dash.cell(row=3, column=c_idx).fill = PatternFill("solid", fgColor="4472C4")
    ws_dash.cell(row=3, column=c_idx).font = HEADER_FONT
    ws_dash.cell(row=3, column=c_idx).alignment = CENTER
    ws_dash.cell(row=3, column=c_idx).border = BORDER_ALL

for i, wk_num in enumerate(sorted(week_to_pillar.keys())):
    r_ = 4 + i
    pillar = week_to_pillar[wk_num]
    is_buf = pillar == BUFFER_SHEET_KEY
    title = BUFFER_TITLE if is_buf else PILLAR_TITLES[pillar]
    light = BUFFER_LIGHT if is_buf else PILLAR_COLORS[pillar]

    ws_dash.cell(row=r_, column=1, value=f"Tuần {wk_num}").border = BORDER_ALL
    ws_dash.cell(row=r_, column=1).alignment = CENTER
    ws_dash.cell(row=r_, column=2, value=title).border = BORDER_ALL
    ws_dash.cell(row=r_, column=2).fill = PatternFill("solid", fgColor=light)
    count_formula = (
        f"=COUNTIFS('{pillar}'!$M${FIRST_DATA_ROW}:$M${LAST_DATA_ROW},{wk_num},"
        f"'{pillar}'!$B${FIRST_DATA_ROW}:$B${LAST_DATA_ROW},\"<>\")"
    )
    ws_dash.cell(row=r_, column=3, value=count_formula).border = BORDER_ALL
    ws_dash.cell(row=r_, column=3).alignment = CENTER
    if is_buf:
        ws_dash.cell(row=r_, column=4, value="— (tuần buffer, không bắt buộc)").border = BORDER_ALL
    else:
        ws_dash.cell(row=r_, column=4, value=f'=IF(C{r_}>=3,"✓ Đạt","✗ Cần bổ sung")').border = BORDER_ALL
    ws_dash.cell(row=r_, column=4).alignment = CENTER

last_wk_row = 3 + num_weeks
ws_dash.conditional_formatting.add(f"D4:D{last_wk_row}",
    FormulaRule(formula=['D4="✓ Đạt"'], fill=cf_fill("C6EFCE")))

dash2_row = last_wk_row + 2
ws_dash.cell(row=dash2_row, column=1, value="Loại 3H").font = HEADER_FONT
ws_dash.cell(row=dash2_row, column=2, value="Số lượng").font = HEADER_FONT
ws_dash.cell(row=dash2_row, column=3, value="Tỷ lệ %").font = HEADER_FONT
ws_dash.cell(row=dash2_row, column=4, value="Baseline đề xuất").font = HEADER_FONT
ws_dash.cell(row=dash2_row, column=5, value="Đánh giá").font = HEADER_FONT
for c_idx in (1, 2, 3, 4, 5):
    ws_dash.cell(row=dash2_row, column=c_idx).fill = PatternFill("solid", fgColor="4472C4")
    ws_dash.cell(row=dash2_row, column=c_idx).font = HEADER_FONT
    ws_dash.cell(row=dash2_row, column=c_idx).alignment = CENTER
    ws_dash.cell(row=dash2_row, column=c_idx).border = BORDER_ALL

h3_baseline = CONFIG["h3_baseline"]
h3_count_rows = {}
for i, h3 in enumerate(H3_LIST):
    r_ = dash2_row + 1 + i
    h3_count_rows[h3] = r_
    ws_dash.cell(row=r_, column=1, value=h3).border = BORDER_ALL
    ws_dash.cell(row=r_, column=1).fill = PatternFill("solid", fgColor=H3_FILL[h3])
    ws_dash.cell(row=r_, column=1).font = Font(bold=True, color=H3_COLORS[h3])

    count_parts = [f"COUNTIF('{sk}'!$C${FIRST_DATA_ROW}:$C${LAST_DATA_ROW},\"{h3}\")" for sk in all_pillar_sheet_keys]
    ws_dash.cell(row=r_, column=2, value="=" + "+".join(count_parts)).border = BORDER_ALL
    ws_dash.cell(row=r_, column=2).alignment = CENTER

for i, h3 in enumerate(H3_LIST):
    r_ = h3_count_rows[h3]
    total_ref_parts = [f"COUNTIF('{sk}'!$B${FIRST_DATA_ROW}:$B${LAST_DATA_ROW},\"<>\")" for sk in all_pillar_sheet_keys]
    total_formula = "(" + "+".join(total_ref_parts) + ")"
    ws_dash.cell(row=r_, column=3, value=f"=IF({total_formula}=0,0,B{r_}/{total_formula})").border = BORDER_ALL
    ws_dash.cell(row=r_, column=3).number_format = "0.0%"
    ws_dash.cell(row=r_, column=3).alignment = CENTER

    lo, hi = h3_baseline[h3]
    ws_dash.cell(row=r_, column=4, value=f"{lo:.0%} - {hi:.0%}").border = BORDER_ALL
    ws_dash.cell(row=r_, column=4).alignment = CENTER

    eval_formula = f'=IF(AND(C{r_}>={lo},C{r_}<={hi}),"✓ Đạt baseline",IF(C{r_}<{lo},"✗ Thấp hơn baseline","✗ Cao hơn baseline"))'
    ws_dash.cell(row=r_, column=5, value=eval_formula).border = BORDER_ALL
    ws_dash.cell(row=r_, column=5).alignment = CENTER

last_h3_row = dash2_row + len(H3_LIST)
ws_dash.conditional_formatting.add(f"E{dash2_row+1}:E{last_h3_row}",
    FormulaRule(formula=[f'LEFT(E{dash2_row+1},1)="✓"'], fill=cf_fill("C6EFCE")))
ws_dash.conditional_formatting.add(f"E{dash2_row+1}:E{last_h3_row}",
    CellIsRule(operator="equal", formula=['"✗ Thấp hơn baseline"'], fill=cf_fill("FFC7CE")))
ws_dash.conditional_formatting.add(f"E{dash2_row+1}:E{last_h3_row}",
    CellIsRule(operator="equal", formula=['"✗ Cao hơn baseline"'], fill=cf_fill("FCE4D6")))

total_row = last_h3_row + 2
total_all_formula = "=" + "+".join(
    f"COUNTIF('{sk}'!$B${FIRST_DATA_ROW}:$B${LAST_DATA_ROW},\"<>\")" for sk in all_pillar_sheet_keys)
ws_dash.cell(row=total_row, column=1, value="Tổng số Content Cluster trong tháng:").font = Font(bold=True, italic=True)
ws_dash.cell(row=total_row, column=2, value=total_all_formula).font = Font(bold=True)
ws_dash.cell(row=total_row + 1, column=1, value="Số tuần lịch trong tháng:").font = Font(bold=True, italic=True)
ws_dash.cell(row=total_row + 1, column=2, value=num_weeks).font = Font(bold=True)
ws_dash.cell(row=total_row + 2, column=1, value="Buffer newsjacking mục tiêu (Config):").font = Font(bold=True, italic=True)
ws_dash.cell(row=total_row + 2, column=2, value=f"=Config!$B${CFG_ROW['buffer_pct']}").font = Font(bold=True)
ws_dash.cell(row=total_row + 2, column=2).number_format = "0%"

autofit(ws_dash, [16, 42, 16, 20, 20])

# ============================================================================
# 7. SHEET HƯỚNG DẪN SỬ DỤNG — bao gồm cách tái sử dụng cho ngành khác
# ============================================================================

ws_guide = wb.create_sheet("Huong Dan", 2)
ws_guide.sheet_view.showGridLines = False
ws_guide.column_dimensions["A"].width = 118

guide_lines = [
    ("HƯỚNG DẪN SỬ DỤNG & TÁI SỬ DỤNG TEMPLATE CONTENT CALENDAR ĐA NGÀNH", True, 14, "203864"),
    ("", False, 10, None),
    ("0. FILE NÀY LÀ TEMPLATE TRỐNG — chưa có nội dung cluster mẫu nào:", True, 11, "C00000"),
    ("   - Tất cả sheet Pillar (và Tuan Buffer) hiện KHÔNG có dòng dữ liệu nào — chỉ có cấu trúc cột, header,", False, 10, None),
    ("     dropdown (Loại 3H/Kênh/Trạng thái/Phễu), công thức Tuần/Deadline, và định dạng màu theo trạng thái.", False, 10, None),
    ("   - Content Calendar và Dashboard vẫn hiển thị đúng lưới/công thức nhưng sẽ trống/= 0 cho tới khi bạn điền dữ liệu.", False, 10, None),
    ("   - Cách thêm nội dung thật (khuyến nghị — KHÔNG cần chạy lại script): mở sheet Pillar tương ứng, điền trực tiếp", False, 10, None),
    ("     vào các cột: Mã Cluster, Tên bài/Chủ đề, Loại 3H (chọn từ dropdown), Mục tiêu phễu, Persona, Từ khóa, Tone,", False, 10, None),
    ("     Độ dài, Outline, CTA, Kênh (dropdown), Ngày đăng (chọn ngày — cột 'Tuần (auto)' và 'Deadline' tự tính),", False, 10, None),
    ("     Người phụ trách, Trạng thái (dropdown), KPI dự kiến. Content Calendar sẽ tự hiển thị ngay khi Excel tính lại.", False, 10, None),
    ("   - Cách thêm nội dung bằng script (khi cần tạo lại file từ đầu với dữ liệu khác): điền dữ liệu vào", False, 10, None),
    ("     PILLAR_DATA_RAW ở đầu file generate-content-calendar.py theo đúng khóa từng Pillar, rồi chạy lại script —", False, 10, None),
    ("     ngày đăng và vị trí buffer sẽ được TỰ ĐỘNG tính toán và phân bổ xen kẽ (xem mục 5).", False, 10, None),
    ("", False, 10, None),
    ("1. KIẾN TRÚC TEMPLATE (vì sao tái sử dụng được cho ngành khác):", True, 11, "C00000"),
    ("   - Sheet 'Config': toàn bộ thông tin đặc thù (brand, persona, năm/tháng, danh sách Pillar, kênh, trạng thái,", False, 10, None),
    ("     baseline tỷ lệ 3H, % buffer newsjacking) được tách riêng khỏi logic. Các sheet khác THAM CHIẾU vào Config", False, 10, None),
    ("     bằng công thức (=Config!$B$5, =Config!$B$11...) nên đổi Config sẽ tự cập nhật tiêu đề/công thức liên quan.", False, 10, None),
    ("   - Muốn áp dụng cho ngành khác (ví dụ mỹ phẩm, giáo dục, B2B SaaS...): mở file script generate-content-calendar.py,", False, 10, None),
    ("     sửa khối CONFIG và danh sách PILLARS ở đầu file (tên brand, persona, tên/màu Pillar), điền dữ liệu vào", False, 10, None),
    ("     PILLAR_DATA_RAW theo nội dung cluster của ngành mới, rồi chạy lại script — toàn bộ sheet/công thức tự sinh lại đúng cấu trúc.", False, 10, None),
    ("   - Nếu chỉ cần sửa nội dung/ngày đăng (không đổi cấu trúc Pillar) thì KHÔNG cần chạy lại script — sửa trực tiếp", False, 10, None),
    ("     trong Excel ở các sheet Pillar là đủ (xem mục 3 dưới).", False, 10, None),
    ("", False, 10, None),
    ("2. Sheet 'Content Calendar': Lưới theo TUẦN (hàng) x THỨ TRONG TUẦN (cột: Thứ 2 -> CN).", True, 11, None),
    ("   - Mỗi ô nội dung là CÔNG THỨC TEXTJOIN + IF (không phải giá trị tĩnh) quét toàn bộ sheet Pillar + Tuần Buffer", False, 10, None),
    ("     để tìm cluster có 'Ngày đăng' khớp đúng ngày đó. Công thức chỉ dùng TEXTJOIN/IF (Excel 2016/2019+),", False, 10, None),
    ("     KHÔNG dùng FILTER/XLOOKUP (chỉ có ở Excel 365) để đảm bảo tương thích các phiên bản Excel cũ hơn.", False, 10, None),
    ("   - Mỗi ô hiển thị nội dung là công thức NỐI CHUỖI đơn giản (=TRIM(helper1&helper2&...)), tham chiếu tới các", False, 10, None),
    ("     'ô helper' ẩn phía bên phải lưới — mỗi helper là 1 công thức TEXTJOIN+IF riêng cho TỪNG sheet Pillar", False, 10, None),
    ("     (giữ mỗi công thức ngắn, dễ kiểm tra, tương thích tối đa). KHÔNG xóa vùng cột ẩn này (đã Hide) vì Calendar phụ thuộc vào đó.", False, 10, None),
    ("   - Click vào ô nội dung để nhảy tới dòng chi tiết trong sheet Pillar tương ứng (hyperlink gợi ý, không phải nội dung động).", False, 10, None),
    ("", False, 10, None),
    ("3. LIÊN KẾT ĐỘNG CALENDAR <-> PILLAR (quan trọng nhất):", True, 11, "C00000"),
    ("   - Sheet Pillar là NGUỒN DỮ LIỆU DUY NHẤT. Cột 'Ngày đăng' (L) là ô nhập trực tiếp — sửa ngày, tên bài, loại 3H,", False, 10, None),
    ("     trạng thái... ngay trong sheet Pillar, sheet Content Calendar và Dashboard SẼ TỰ TÍNH LẠI ngay lập tức (Excel tự", False, 10, None),
    ("     recalculate công thức) — không cần chạy lại script Python.", False, 10, None),
    ("   - XÓA một dòng cluster (hoặc xóa nội dung ô 'Tên bài/Chủ đề') sẽ khiến ô Calendar tương ứng ngày đó tự trống lại.", False, 10, None),
    ("   - THÊM cluster mới: điền vào một dòng còn trống trong phạm vi đã kẻ sẵn của sheet Pillar (mỗi sheet dự trù", False, 10, None),
    (f"     {MAX_ROWS_PER_PILLAR} dòng, dòng dữ liệu từ hàng {FIRST_DATA_ROW} đến {LAST_DATA_ROW}), điền Ngày đăng — Calendar tự nhận diện,", False, 10, None),
    ("     KHÔNG cần sửa công thức. Nếu cần nhiều hơn số dòng dự trù, phải mở lại script, tăng MAX_ROWS_PER_PILLAR và chạy lại.", False, 10, None),
    ("   - Cột 'Tuần (auto)' (M) là công thức tự tính tuần lịch từ Ngày đăng + ngày đầu tháng ở Config — không nhập tay.", False, 10, None),
    ("   - Cột 'Ngày đăng' (L) hiển thị và cần nhập theo dd/mm/yyyy (đã đặt number format + Data Validation kiểu Date với", False, 10, None),
    ("     gợi ý khi click vào ô). Excel diễn giải chuỗi ngày bạn gõ tay dựa theo locale hệ thống máy — nếu máy đang đặt", False, 10, None),
    ("     locale tiếng Anh (US), gõ '25/12/2026' có thể bị hiểu nhầm. Cách an toàn nhất: click vào ô và dùng lịch chọn", False, 10, None),
    ("     ngày (date picker) của Excel thay vì gõ tay, hoặc gõ theo thứ tự năm-tháng-ngày (2026-12-25) để tránh nhầm lẫn.", False, 10, None),
    ("", False, 10, None),
    ("4. Sheet 'Tuan Buffer' và nguyên tắc xử lý TUẦN DƯ:", True, 11, "C00000"),
    ("   - Nếu số tuần lịch trong tháng NHIỀU HƠN số Content Pillar, các tuần dư (luôn ở cuối tháng) KHÔNG bị gán quay", False, 10, None),
    ("     lại Pillar 1 (tránh làm Pillar 1 bị trải ra 2 tuần cách xa nhau trong tháng, gây mất tập trung chủ đề tuần).", False, 10, None),
    ("   - Thay vào đó, các tuần dư trở thành 'Tuần Buffer' — sheet riêng dùng để chèn nội dung newsjacking phát sinh", False, 10, None),
    ("     hoặc tổng hợp bài xen ngang từ nhiều Pillar khi cần. Ô Calendar của tuần buffer hiển thị màu xám nếu chưa có nội dung.", False, 10, None),
    ("", False, 10, None),
    ("5. Buffer newsjacking XEN KẼ trong tuần (mục 2.5 tài liệu Module 2 — 'chừa buffer là yêu cầu bắt buộc'):", True, 11, "C00000"),
    (f"   - Khi phân bổ ngày đăng cho từng Pillar (chạy script), thuật toán CHỌN VỊ TRÍ BUFFER TRƯỚC bằng cách", False, 10, None),
    (f"     trải đều {BUFFER_PCT:.0%} số ngày trong tuần theo bước nhảy cố định (hàm pick_evenly_spaced) — nên ngày buffer", False, 10, None),
    ("     xen kẽ rải rác giữa tuần (đầu/giữa/cuối), KHÔNG dồn cố định vào cuối tuần như cách làm cũ. Nhờ vậy buffer có thể", False, 10, None),
    ("     dùng phản ứng với sự kiện thời sự phát sinh bất kỳ ngày nào trong tuần, không chỉ cuối tuần.", False, 10, None),
    ("   - Các cluster (khi được điền) sau đó lấp vào đúng những ngày CÒN LẠI (không phải ngày buffer), cũng được trải", False, 10, None),
    ("     đều bằng thuật toán tương tự để tránh dồn cụm liên tiếp.", False, 10, None),
    ("   - Có thể chỉnh tỷ lệ buffer ở CONFIG['buffer_pct'] trong script (khung đề xuất 10-15%).", False, 10, None),
    ("", False, 10, None),
    ("6. Sheet 'Dashboard Can Bang':", True, 11, None),
    ("   - Bảng 1: đếm số cluster mỗi Tuần bằng COUNTIFS tham chiếu trực tiếp cột 'Tuần (auto)' của sheet Pillar tương ứng", False, 10, None),
    ("     -> tự cập nhật khi sửa/thêm/xóa dòng, đối chiếu ngưỡng tối thiểu 3 bài/tuần (trừ tuần buffer).", False, 10, None),
    ("   - Bảng 2: tỷ lệ % Hero/Hub/Hygiene toàn tháng bằng COUNTIF trên TẤT CẢ sheet Pillar + Buffer, có cột 'Đánh giá'", False, 10, None),
    ("     tự so sánh với baseline lấy từ CONFIG['h3_baseline'] (Hero 5-10%, Hub 40-50%, Hygiene 40-55%) và báo", False, 10, None),
    ("     '✓ Đạt baseline' / '✗ Thấp hơn baseline' / '✗ Cao hơn baseline' tương ứng.", False, 10, None),
    ("   - File chưa có dữ liệu cluster nên Bảng 1 sẽ hiện 0 và Bảng 2 sẽ hiện 0.0% ở mọi dòng — đây là kết quả ĐÚNG cho", False, 10, None),
    ("     một template trống, không phải lỗi. Dashboard sẽ tự cập nhật khi bạn điền dữ liệu vào các sheet Pillar.", False, 10, None),
    ("", False, 10, None),
    ("7. Nguồn tham chiếu chiến lược: Module 2 - Chiến lược Nội dung (calendar-3H.md) - Content Pillar, Topic Cluster,", True, 11, None),
    ("   Mô hình 3H (Hero-Hub-Hygiene), Content Calendar, Content Brief. Cấu trúc cột Pillar/Content Brief trong file này", False, 10, None),
    ("   được xây dựng theo đúng khung lý thuyết ở tài liệu Module 2 — điền tên Pillar, brand, persona thật vào", False, 10, None),
    ("   CONFIG/PILLARS trong script (hoặc trực tiếp trong sheet Config/Pillar) khi áp dụng cho thương hiệu cụ thể.", False, 10, None),
    ("", False, 10, None),
    ("8. VÍ DỤ MINH HỌA 1 CONTENT CLUSTER ĐẦY ĐỦ (chỉ mang tính hướng dẫn — KHÔNG phải dữ liệu thật trong sheet Pillar):", True, 11, "C00000"),
    ("   Bảng dưới đây minh họa cách điền đầy đủ 17 trường của MỘT dòng cluster trong sheet Pillar, giúp hình dung rõ", False, 10, None),
    ("   ý nghĩa từng cột trước khi tự điền dữ liệu thật. Xem bảng ở các dòng ngay dưới đây.", False, 10, None),
]
for i, (text, bold, size, color) in enumerate(guide_lines, start=1):
    cell = ws_guide.cell(row=i, column=1, value=text)
    cell.font = Font(bold=bold, size=size, color=color if color else "000000")
    cell.alignment = Alignment(wrap_text=True, vertical="top")

# ----------------------------------------------------------------------------
# 7.1 Bảng ví dụ minh họa 1 cluster mẫu đầy đủ (hiển thị dạng Trường | Giá trị
#     ví dụ | Giải thích) — chỉ để hướng dẫn, KHÔNG phải dữ liệu thật.
# ----------------------------------------------------------------------------

example_cluster_row = len(guide_lines) + 2
example_table_header = ["Trường (cột trong sheet Pillar)", "Giá trị ví dụ minh họa", "Giải thích"]
example_cluster_fields = [
    ("Mã Cluster", "P1-01", "Mã tự đặt để dễ tra cứu, gợi ý: <tên Pillar>-<số thứ tự>."),
    ("Tên bài/Chủ đề", "5 dấu hiệu nhận biết [sản phẩm] không đạt chuẩn chất lượng",
     "Tên bài viết/nội dung cụ thể, nên chứa từ khóa chính."),
    ("Loại 3H", "Hygiene", "Chọn 1 trong 3: Hero (hiếm, đầu tư cao) / Hub (định kỳ) / Hygiene (liên tục, SEO)."),
    ("Mục tiêu phễu", "TOFU", "TOFU (nhận biết) / MOFU (cân nhắc) / BOFU (quyết định mua)."),
    ("Persona mục tiêu", "[Điền persona chính đã xây dựng ở Module 2]", "Đối tượng cụ thể bài viết hướng tới."),
    ("Từ khóa chính", "cách nhận biết [sản phẩm] kém chất lượng, mẹo chọn [sản phẩm]",
     "Từ khóa SEO chính + 1-2 từ khóa dài liên quan."),
    ("Tone giọng điệu", "Gần gũi, đáng tin cậy, có dẫn chứng cụ thể", "Văn phong khi viết, khớp với Brand Voice đã định nghĩa."),
    ("Độ dài mong muốn", "600-800 từ", "Ước lượng độ dài bài viết hoặc thời lượng video."),
    ("Outline/Dàn ý chính", "Hook mở bài -> 5 dấu hiệu nhận biết (mỗi dấu hiệu 1 đoạn) -> Lời khuyên/CTA",
     "Dàn ý ngắn gọn để người viết bám theo, tránh lệch trọng tâm."),
    ("CTA mong muốn", "Lưu bài để tham khảo khi mua hàng lần tới", "Hành động cụ thể muốn người đọc thực hiện."),
    ("Kênh đăng", "Facebook", "Chọn từ dropdown: Facebook/Zalo OA/Website/TikTok/Email (tùy Config)."),
    ("Ngày đăng", "01/09/2026", "Chọn ngày cụ thể — cột Tuần (auto) và Deadline sẽ tự tính theo ngày này."),
    ("Tuần (auto)", "1 (tự động)", "KHÔNG nhập tay — công thức tự tính tuần lịch chứa Ngày đăng."),
    ("Deadline", "30/08/2026 (tự động)", "KHÔNG nhập tay — mặc định = Ngày đăng trừ 2 ngày để có thời gian duyệt."),
    ("Người phụ trách", "Writer A", "Người chịu trách nhiệm sản xuất nội dung này."),
    ("Trạng thái", "Draft", "Draft/Review/Approved/Published — đổi màu tự động theo trạng thái."),
    ("KPI dự kiến", "Organic traffic; Time on page > 1'30\"", "Chỉ số đo hiệu quả, nên khác nhau theo từng loại 3H (xem mục 6)."),
]
ws_guide.cell(row=example_cluster_row - 1, column=1,
    value="Bảng ví dụ (17 trường của 1 dòng cluster mẫu):").font = Font(bold=True, size=10, color="203864")
for c_idx, h in enumerate(example_table_header, start=1):
    cell = ws_guide.cell(row=example_cluster_row, column=c_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.border = BORDER_ALL
    cell.alignment = CENTER
ws_guide.column_dimensions["B"].width = 55
ws_guide.column_dimensions["C"].width = 60
for i, (field, value, note) in enumerate(example_cluster_fields, start=1):
    r_ = example_cluster_row + i
    ws_guide.cell(row=r_, column=1, value=field).font = Font(bold=True, size=9)
    ws_guide.cell(row=r_, column=2, value=value).font = Font(size=9)
    ws_guide.cell(row=r_, column=3, value=note).font = Font(italic=True, size=9, color="595959")
    for c_idx in (1, 2, 3):
        cell = ws_guide.cell(row=r_, column=c_idx)
        cell.border = BORDER_ALL
        cell.alignment = WRAP_TOP
        cell.fill = PatternFill("solid", fgColor="F2F2F2" if i % 2 == 0 else "FFFFFF")

# ============================================================================
# 8. THỨ TỰ SHEET & LƯU FILE
# ============================================================================

desired_order = ["Huong Dan", "Content Calendar", "Dashboard Can Bang", "Config"] + all_pillar_sheet_keys
wb._sheets = [wb[name] for name in desired_order]
wb.active = 0

# QUAN TRỌNG: buộc Excel TÍNH LẠI TOÀN BỘ công thức mỗi khi mở file. Nếu không có
# cờ này, các ô công thức (đặc biệt các 'helper' TEXTJOIN+IF ẩn của Content Calendar)
# sẽ hiển thị giá trị cache RỖNG (do openpyxl không tự tính formula khi ghi file),
# khiến người dùng nhìn thấy sửa "Tên bài" trong sheet Pillar nhưng Calendar không
# tự hiển thị nội dung mới cho tới khi tự ép Excel tính lại (F9). Đặt fullCalcOnLoad
# = True để Excel luôn tính lại ngay khi mở, không phụ thuộc vào giá trị cache.
wb.calculation.fullCalcOnLoad = True
wb.calculation.calcMode = "auto"

wb.save(OUT_PATH)
print(f"Đã lưu file: {OUT_PATH}")
