#!/usr/bin/env python3
"""
Generate multi-platform short video content analysis Excel file.
Applies to TikTok, Facebook video formats and related short-form platforms.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
import os

OUTPUT_PATH = "/Users/phamvannam/Documents/GitHub/hello-affiliate/content-marketing/code/TikTok_Video_Analysis.xlsx"

# ─── COLOR PALETTE ────────────────────────────────────────────────────────────
C_HEADER_DARK   = "1A3C5E"   # deep navy
C_HEADER_MID    = "2E7D52"   # forest green (vegan)
C_HEADER_LIGHT  = "4CAF82"   # mint green
C_SECTION_1     = "E8F5E9"   # very light green
C_SECTION_2     = "E3F2FD"   # very light blue
C_SECTION_3     = "FFF9C4"   # very light yellow
C_SECTION_4     = "FCE4EC"   # very light pink
C_SECTION_5     = "F3E5F5"   # very light purple
C_ACCENT        = "FF6F00"   # amber
C_WHITE         = "FFFFFF"
C_LIGHT_GRAY    = "F5F5F5"
C_BORDER        = "BDBDBD"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def border_thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_cell(ws, row, col, value, bold=False, bg=None, font_color="000000",
             font_size=10, align=None, italic=False, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, size=font_size, color=font_color, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.border = border_thin()
    cell.alignment = align or left_wrap()
    if num_format:
        cell.number_format = num_format
    return cell

def merge_header(ws, row, col_start, col_end, value, bg, font_color="FFFFFF",
                 font_size=11):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font(bold=True, size=font_size, color=font_color)
    cell.fill = fill(bg)
    cell.alignment = center()
    cell.border = border_thin()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: HƯỚNG DẪN SỬ DỤNG
# ══════════════════════════════════════════════════════════════════════════════
def create_guide_sheet(wb):
    ws = wb.create_sheet("📖 Hướng dẫn sử dụng", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24

    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "📊 KHUNG PHÂN TÍCH CONTENT VIDEO ĐA NỀN TẢNG: TIKTOK + FACEBOOK"
    c.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    c.fill = fill(C_HEADER_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 42

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "Áp dụng cho TikTok, Facebook Reels, Facebook Feed Video, Facebook Watch, Instagram Reels và các biến thể short-form liên quan"
    c.font = Font(bold=False, size=11, color="FFFFFF", name="Calibri", italic=True)
    c.fill = fill(C_HEADER_MID)
    c.alignment = center()
    ws.row_dimensions[2].height = 32

    def section_title(row, title, bg):
        ws.merge_cells(f"B{row}:E{row}")
        c = ws[f"B{row}"]
        c.value = title
        c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        c.fill = fill(bg)
        c.alignment = left_wrap()
        ws.row_dimensions[row].height = 24

    def row_item(row, label, desc, note=""):
        ws.row_dimensions[row].height = 42
        set_cell(ws, row, 2, label, bold=True, bg=C_LIGHT_GRAY, font_size=10)
        set_cell(ws, row, 3, desc, bold=False, bg=C_WHITE, font_size=10)
        set_cell(ws, row, 4, note, bold=False, bg=C_WHITE, font_size=9, italic=True,
                 font_color="666666")

    section_title(4, "🗂️  CÁC SHEET TRONG FILE", C_HEADER_DARK)
    for ci, h in enumerate(["Sheet", "Mô tả", "Ghi chú"], 2):
        set_cell(ws, 5, ci, h, bold=True, bg=C_HEADER_MID, font_color="FFFFFF",
                 align=center())
    sheets_info = [
        ("📖 Hướng dẫn sử dụng", "Giải thích khung đa nền tảng, nguồn dữ liệu và quy tắc chấm điểm", "Không chỉnh sửa"),
        ("📋 Dữ liệu Video", "Nhập metadata, nền tảng và số liệu hiệu suất của từng video", "Sheet nhập liệu chính"),
        ("🎬 Phân tích Phân cảnh", "Phân tích Hook / Thân / Kết cho từng video", "Nhập sau khi xem video"),
        ("⭐ Bảng Điểm Tổng hợp", "Tổng hợp điểm và phục vụ so sánh chéo video", "Tự động tính"),
        ("📐 Thang điểm & Trọng số", "Giữ benchmark TikTok hiện có và bổ sung benchmark Facebook tham khảo", "Tham khảo khi chấm điểm"),
    ]
    for ri, (sname, desc, note) in enumerate(sheets_info, 6):
        row_item(ri, sname, desc, note)

    section_title(12, "📌  NHÓM TIÊU CHÍ & TRỌNG SỐ", C_HEADER_MID)
    for ci, h in enumerate(["Nhóm tiêu chí", "Mô tả", "Trọng số (%)"], 2):
        set_cell(ws, 13, ci, h, bold=True, bg=C_HEADER_MID, font_color="FFFFFF",
                 align=center())
    groups = [
        ("A. Thông tin định danh", "STT, link, ngày đăng, nền tảng, mùa lễ hội, loại quảng cáo", "—"),
        ("B. Metadata Video", "Thời lượng, thể loại, sản phẩm, CTA, bối cảnh phân phối", "—"),
        ("C. Chỉ số tương tác (Engagement)", "Views, likes/reactions, comments, shares, saves, completion/watch rate, CTR", "35%"),
        ("D. Chỉ số tăng trưởng (Growth)", "Follow mới, điểm bắt trend", "10%"),
        ("E. Chất lượng nội dung – Hook", "Giữ nguyên cấu trúc Hook; với Facebook có thể đánh giá trong 3–5 giây đầu", "25%"),
        ("F. Chất lượng nội dung – Thân", "Luồng lập luận, bằng chứng, social proof, điểm gãy", "20%"),
        ("G. Chất lượng nội dung – Kết/CTA", "Chốt thông điệp và lời kêu gọi hành động rõ ràng", "10%"),
        ("H. Điểm tổng hợp (Composite)", "Giữ nguyên công thức hiện có để so sánh nhất quán", "100%"),
    ]
    bg_list = [C_SECTION_1, C_SECTION_1, C_SECTION_2, C_SECTION_3,
               C_SECTION_4, C_SECTION_5, C_SECTION_1, C_LIGHT_GRAY]
    for ri, ((grp, desc, wt), bg) in enumerate(zip(groups, bg_list), 14):
        ws.row_dimensions[ri].height = 38
        set_cell(ws, ri, 2, grp, bold=True, bg=bg, font_size=10)
        set_cell(ws, ri, 3, desc, bg=C_WHITE, font_size=10)
        set_cell(ws, ri, 4, wt, bg=C_WHITE, font_size=10, align=center(),
                 bold=(wt != "—"))

    section_title(23, "🔎  NGUỒN SỐ LIỆU THEO NỀN TẢNG", C_ACCENT)
    sources = [
        ("TikTok", "TikTok Studio / TikTok Creative Center / TikTok Shop Analytics", "Ưu tiên dùng Completion Rate, CTR giỏ/link, Saves"),
        ("Facebook", "Business Suite / Meta Ads Manager / Facebook Page Insights", "Đối chiếu ThruPlay, Average Watch Time, 3-second video views, Link CTR"),
        ("Lưu ý", "Facebook có Reactions phân loại Love/Haha/Wow/Sad/Angry; TikTok không có", "Reactions chi tiết dùng để chẩn đoán cảm xúc, chưa cộng riêng vào composite"),
    ]
    for ri, (label, desc, note) in enumerate(sources, 24):
        row_item(ri, label, desc, note)

    section_title(28, "✍️  HƯỚNG DẪN NHẬP LIỆU", C_ACCENT)
    steps = [
        ("Bước 1", "Xem video và ghi lại URL/link, chọn đúng nền tảng tại sheet '📋 Dữ liệu Video'"),
        ("Bước 2", "Điền metadata và chỉ số từ nguồn phù hợp: TikTok Studio/TikTok Creative Center hoặc Business Suite/Meta Ads Manager/Facebook Insights"),
        ("Bước 3", "Với Facebook, quy đổi chỉ số tương đương trước khi nhập: ThruPlay hoặc watch rate dùng để tham chiếu Completion Rate; Link CTR dùng cho CTR"),
        ("Bước 4", "Xem lại video, phân tích theo 3 giai đoạn Hook / Thân / Kết tại sheet '🎬 Phân tích Phân cảnh'"),
        ("Bước 5", "Sheet '⭐ Bảng Điểm Tổng hợp' dùng cùng khung composite để so sánh chéo video"),
        ("Mẹo", "Luôn lọc theo nền tảng trước khi so benchmark; benchmark Facebook trong file là ước lượng tham khảo theo thông lệ ngành"),
    ]
    for ri, (step, desc) in enumerate(steps, 29):
        ws.row_dimensions[ri].height = 34
        set_cell(ws, ri, 2, step, bold=True, bg=C_SECTION_3, font_size=10)
        set_cell(ws, ri, 3, desc, bg=C_WHITE, font_size=10)

    section_title(36, "📐  CÔNG THỨC TÍNH ĐIỂM & TƯƠNG QUAN", C_HEADER_MID)
    formulas = [
        ("Engagement Rate (ER)", "= (Likes/Reaction + Comments + Shares + Saves) / Views × 100%",
         "TikTok thường cao hơn Facebook; xem thêm benchmark riêng tại sheet '📐 Thang điểm & Trọng số'"),
        ("Completion / Watch Rate", "= Lượt xem hết hoặc chỉ số tương đương / Tổng lượt xem × 100%",
         "Facebook có thể tham chiếu ThruPlay, Average Watch Time, 3-second video views tùy nguồn"),
        ("CTR CTA / Giỏ hàng", "= Lượt nhấp CTA hoặc link / Tổng lượt xem × 100%",
         "Dùng Link CTR hoặc outbound CTR nếu không có TikTok Shop"),
        ("Điểm Nội dung", "= Hook×0.45 + Thân×0.35 + Kết×0.20",
         "Giữ nguyên thang 1–10"),
        ("Điểm Tổng (Composite)", "= ER×0.35 + CompRate×0.15 + CTR×0.15 + DiemNoiDung×0.25 + GrowthScore×0.10",
         "Giữ nguyên cấu trúc TikTok để so sánh nhất quán"),
        ("Hệ số tương quan Pearson", "Dùng =CORREL(range1, range2) để so sánh Hook, watch rate, CTR theo từng nền tảng hoặc giữa 2 nền tảng",
         "Nên lọc riêng TikTok và Facebook trước khi kết luận"),
    ]
    for ci, h in enumerate(["Công thức", "Cách tính", "Ngưỡng / Ghi chú"], 2):
        set_cell(ws, 37, ci, h, bold=True, bg=C_HEADER_MID, font_color="FFFFFF",
                 align=center())
    for ri, (name, formula, note) in enumerate(formulas, 38):
        ws.row_dimensions[ri].height = 38
        set_cell(ws, ri, 2, name, bold=True, bg=C_SECTION_2, font_size=10)
        set_cell(ws, ri, 3, formula, bg=C_WHITE, font_size=10, font_color="1A237E")
        set_cell(ws, ri, 4, note, bg=C_SECTION_1, font_size=9, italic=True)

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: DỮ LIỆU VIDEO
# ══════════════════════════════════════════════════════════════════════════════
def create_data_sheet(wb):
    ws = wb.create_sheet("📋 Dữ liệu Video")
    ws.freeze_panes = "D3"
    ws.sheet_view.showGridLines = False

    # ── ROW 1: Section headers (merged) ──
    sections = [
        (1, 3,  "A. THÔNG TIN ĐỊNH DANH",           C_HEADER_DARK,  3),
        (4, 10, "B. METADATA VIDEO",                 C_HEADER_MID,   7),
        (11,19, "C. CHỈ SỐ TƯƠNG TÁC (ENGAGEMENT)", "1565C0",        9),
        (20,22, "D. TĂNG TRƯỞNG",                   "6A1B9A",        3),
        (23,25, "E. ĐIỂM TÍNH TOÁN",                C_ACCENT,        3),
    ]
    for cs, ce, title, bg, _ in sections:
        merge_header(ws, 1, cs, ce, title, bg)

    # ── ROW 2: Column headers ──
    columns = [
        # A. Thông tin định danh (cols 1–3)
        (1, "STT",              12,  C_HEADER_DARK),
        (2, "Link Video",       35,  C_HEADER_DARK),
        (3, "Ngày đăng",        14,  C_HEADER_DARK),
        # B. Metadata (cols 4–10)
        (4,  "Nền tảng",        13,  C_HEADER_MID),
        (5,  "Mùa / Dịp lễ",   18,  C_HEADER_MID),
        (6,  "Loại quảng cáo", 16,  C_HEADER_MID),
        (7,  "Thời lượng (s)", 14,  C_HEADER_MID),
        (8,  "Sản phẩm",       25,  C_HEADER_MID),
        (9,  "Thể loại nội dung", 20, C_HEADER_MID),
        (10, "CTA chính",       20,  C_HEADER_MID),
        # C. Engagement (cols 11–19)
        (11, "Lượt xem",        14,  "1565C0"),
        (12, "Lượt thích",      14,  "1565C0"),
        (13, "Bình luận",       14,  "1565C0"),
        (14, "Lượt share",      14,  "1565C0"),
        (15, "Lượt lưu",        14,  "1565C0"),
        (16, "Xem hết",         14,  "1565C0"),
        (17, "CTR Giỏ hàng(%)", 16,  "1565C0"),
        (18, "Completion Rate(%)", 16, "1565C0"),
        (19, "ER (%)",          14,  "1565C0"),
        # D. Growth (cols 20–22)
        (20, "Follow mới",      14,  "6A1B9A"),
        (21, "Điểm trend (1-10)", 16, "6A1B9A"),
        (22, "Ghi chú trend",   25,  "6A1B9A"),
        # E. Điểm tính toán (cols 23–25)
        (23, "Điểm Nội dung",   16,  C_ACCENT),
        (24, "Điểm Tổng hợp",   16,  C_ACCENT),
        (25, "Xếp hạng",        12,  C_ACCENT),
    ]

    for col_idx, header, width, bg_color in columns:
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        cell.fill = fill(bg_color)
        cell.alignment = center()
        cell.border = border_thin()

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 35

    # ── DATA ROWS: 3 sample rows ──
    sample_data = [
        [1, "https://tiktok.com/@channel/video/001", "2026-07-01",
         "TikTok", "Thường", "Organic",
         45, "Green Protein Shake", "Demo sản phẩm", "Mua ngay link bio",
         150000, 8500, 320, 1200, 950, 75000, 2.5, 50.0,
         "", 120, 7, "Bắt trend nhạc viral",
         "", "", ""],
        [2, "https://facebook.com/example/reel/002", "2026-07-03",
         "Facebook Reels", "Thường", "Paid",
         60, "Collagen Vegan", "Testimonial khách hàng", "Nhấn xem thêm / Link mua hàng",
         230000, 7200, 640, 980, 510, 69000, 1.1, 30.0,
         "", 210, 6, "Creative chạy paid ổn định trên Meta",
         "", "", ""],
        [3, "https://facebook.com/example/videos/003", "2026-07-05",
         "Facebook Feed Video", "Rằm tháng 7", "Organic",
         42, "Omega-3 Vegan", "So sánh sản phẩm", "Nhắn tin nhận tư vấn",
         89000, 2400, 180, 210, 95, 18000, 0.8, 20.2,
         "", 45, 4, "Không dùng trend, tập trung educational angle",
         "", "", ""],
    ]

    bg_rows = [C_WHITE, C_LIGHT_GRAY, C_WHITE]
    for ri, (row_data, row_bg) in enumerate(zip(sample_data, bg_rows), 3):
        ws.row_dimensions[ri].height = 20
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
            cell.fill = fill(row_bg)
            cell.border = border_thin()
            cell.alignment = left_wrap() if ci in [2, 8, 9, 10, 22] else center()
            cell.font = Font(size=9, name="Calibri")
            if ci in [3]:
                cell.number_format = "YYYY-MM-DD"
            if ci in [17, 18]:
                cell.number_format = "0.00"

    # ── Auto-calculated columns (ER, Điểm Tổng, Xếp hạng) ──
    # Row 3 formulas
    for ri in range(3, 6):  # rows 3-5 (sample + blank rows for user)
        row = ri
        # ER formula (col 19): =(L+M+N+O)/K*100
        if ws.cell(row=row, column=11).value:
            er_cell = ws.cell(row=row, column=19)
            er_cell.value = f"=IF(K{row}>0,(L{row}+M{row}+N{row}+O{row})/K{row}*100,0)"
            er_cell.number_format = "0.00"
            er_cell.font = Font(size=9, name="Calibri", italic=True, color="1A237E")
            er_cell.border = border_thin()
            er_cell.alignment = center()
            # Completion Rate (col 18): =P/K*100
            comp_cell = ws.cell(row=row, column=18)
            comp_cell.value = f"=IF(K{row}>0,P{row}/K{row}*100,0)"
            comp_cell.number_format = "0.00"
            comp_cell.font = Font(size=9, name="Calibri", italic=True, color="1A237E")
            comp_cell.border = border_thin()
            comp_cell.alignment = center()

    # Blank rows for user input (rows 6–52)
    for ri in range(6, 53):
        ws.row_dimensions[ri].height = 20
        row_bg = C_WHITE if ri % 2 == 0 else C_LIGHT_GRAY
        for ci in range(1, 26):
            cell = ws.cell(row=ri, column=ci)
            cell.fill = fill(row_bg)
            cell.border = border_thin()
            # Auto formulas for ER and Completion Rate
            if ci == 18:
                cell.value = f"=IF(K{ri}>0,P{ri}/K{ri}*100,0)"
                cell.number_format = "0.00"
                cell.font = Font(size=9, name="Calibri", italic=True, color="1A237E")
                cell.alignment = center()
            elif ci == 19:
                cell.value = f"=IF(K{ri}>0,(L{ri}+M{ri}+N{ri}+O{ri})/K{ri}*100,0)"
                cell.number_format = "0.00"
                cell.font = Font(size=9, name="Calibri", italic=True, color="1A237E")
                cell.alignment = center()

    # Data validation: Nền tảng
    dv_platform = DataValidation(
        type="list",
        formula1='"TikTok,Facebook Reels,Facebook Feed Video,Facebook Watch,Instagram Reels,YouTube Shorts"',
        showDropDown=False
    )
    dv_platform.sqref = "D3:D52"
    ws.add_data_validation(dv_platform)

    # Data validation: Loại quảng cáo
    dv_adtype = DataValidation(
        type="list",
        formula1='"Organic,Paid,Boosted"',
        showDropDown=False
    )
    dv_adtype.sqref = "F3:F52"
    ws.add_data_validation(dv_adtype)

    # Data validation: Mùa lễ hội
    dv_season = DataValidation(
        type="list",
        formula1='"Thường,Tết Nguyên Đán,Rằm tháng 7,8/3,20/10,Giáng sinh,Valentine,11/11,12/12"',
        showDropDown=False
    )
    dv_season.sqref = "E3:E52"
    ws.add_data_validation(dv_season)

    # Data validation: Thể loại nội dung
    dv_content_type = DataValidation(
        type="list",
        formula1='"Demo sản phẩm,Testimonial,So sánh sản phẩm,Giáo dục/Tips,Câu chuyện,Unboxing,Challenge,Live review"',
        showDropDown=False
    )
    dv_content_type.sqref = "I3:I52"
    ws.add_data_validation(dv_content_type)

    # Conditional formatting: ER column (S = col 19)
    ws.conditional_formatting.add(
        "S3:S52",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FF5252",
            mid_type="num",   mid_value=3,     mid_color="FFEB3B",
            end_type="num",   end_value=10,    end_color="4CAF50"
        )
    )
    # Conditional formatting: CTR (Q = col 17)
    ws.conditional_formatting.add(
        "Q3:Q52",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFCDD2",
            mid_type="num",   mid_value=1.5,   mid_color="FFF9C4",
            end_type="num",   end_value=5,     end_color="C8E6C9"
        )
    )

    # Table
    table = Table(displayName="VideoData", ref="A2:Y52")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True,
                           showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: PHÂN TÍCH PHÂN CẢNH
# ══════════════════════════════════════════════════════════════════════════════
def create_scene_sheet(wb):
    ws = wb.create_sheet("🎬 Phân tích Phân cảnh")
    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False

    # Stage sections with colors
    stages = [
        ("HOOK – 5 GIÂY ĐẦU",      C_SECTION_4,  "E53935",  3, 15),
        ("THÂN BÀI – NỘI DUNG CHÍNH", C_SECTION_2, "1565C0", 16, 26),
        ("KẾT / CTA",               C_SECTION_1,  "2E7D32",  27, 33),
    ]

    # Row 1: Stage section headers
    merge_header(ws, 1, 1, 2, "ID VIDEO", C_HEADER_DARK)
    merge_header(ws, 1, 3, 15, "🎯  HOOK – 5 GIÂY ĐẦU  (Trọng số 45%)", "E53935")
    merge_header(ws, 1, 16, 26, "📣  THÂN BÀI – NỘI DUNG CHÍNH  (Trọng số 35%)", "1565C0")
    merge_header(ws, 1, 27, 33, "✅  KẾT / CTA  (Trọng số 20%)", "2E7D32")

    # Row 2: Column headers
    col_defs = [
        # ID
        (1, "STT",                   8,   C_HEADER_DARK),
        (2, "Link Video (tham chiếu)", 30, C_HEADER_DARK),
        # HOOK cols 3-15
        (3,  "Lời thoại / Voice-over (5s)", 28, "E53935"),
        (4,  "Góc máy",              14, "E53935"),
        (5,  "Hành động chính",      20, "E53935"),
        (6,  "Âm thanh / Nhạc nền",  18, "E53935"),
        (7,  "Diễn viên / Nhân vật", 20, "E53935"),
        (8,  "Bối cảnh / Setting",   20, "E53935"),
        (9,  "Trang phục",           18, "E53935"),
        (10, "Pain point đề cập",    25, "E53935"),
        (11, "Có gây tò mò? (Y/N)",  16, "E53935"),
        (12, "Điểm Hook (1-10)",     15, "E53935"),
        (13, "Lý do điểm Hook",      25, "E53935"),
        (14, "Điểm gãy Hook?",       18, "E53935"),
        (15, "Timestamp điểm gãy",   16, "E53935"),
        # THÂN BÀI cols 16-26
        (16, "Nội dung chính truyền đạt", 28, "1565C0"),
        (17, "Bằng chứng / Social proof", 25, "1565C0"),
        (18, "So sánh sản phẩm (Y/N)", 18, "1565C0"),
        (19, "Cách so sánh",         25, "1565C0"),
        (20, "Điểm gãy thân bài (Y/N)", 18, "1565C0"),
        (21, "Mô tả điểm gãy",       25, "1565C0"),
        (22, "Timestamp điểm gãy",   16, "1565C0"),
        (23, "Âm thanh thân bài",    18, "1565C0"),
        (24, "Góc máy thân bài",     16, "1565C0"),
        (25, "Điểm thân bài (1-10)", 15, "1565C0"),
        (26, "Lý do điểm thân bài",  25, "1565C0"),
        # KẾT / CTA cols 27-33
        (27, "Nội dung chốt vấn đề", 25, "2E7D32"),
        (28, "Loại CTA",             18, "2E7D32"),
        (29, "CTA rõ ràng? (Y/N)",   16, "2E7D32"),
        (30, "Nội dung CTA",         25, "2E7D32"),
        (31, "Điểm CTA (1-10)",      15, "2E7D32"),
        (32, "Điểm KẾT tổng (1-10)", 16, "2E7D32"),
        (33, "Điểm Nội dung (Auto)", 18, "2E7D32"),
    ]

    for col_idx, header, width, bg_color in col_defs:
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=8, color="FFFFFF", name="Calibri")
        cell.fill = fill(bg_color)
        cell.alignment = center()
        cell.border = border_thin()

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 45

    # Sample data rows
    sample_scenes = [
        [1, "Video 001",
         "Bạn có biết protein từ thực vật vừa đủ vừa sạch?",
         "Close-up mặt", "Cầm hộp sản phẩm", "Nhạc lo-fi nhẹ",
         "Cô gái 25t", "Bếp hiện đại", "Áo trắng casual",
         "Thiếu protein khi ăn chay", "Y", 8,
         "Hook tốt vì câu hỏi gây tò mò ngay", "N", "",
         "Giải thích 3 lợi ích chính của sản phẩm",
         "Review của 500 khách hàng", "Y",
         "So sánh với whey protein động vật", "N", "", "",
         "Nhạc tăng dần", "Mid-shot", 7, "Thông tin đủ nhưng hơi dài",
         "Sản phẩm giải quyết đúng pain point", "Swipe up mua ngay",
         "Y", "Link bio – giảm 20% hôm nay", 9, 8, None],
    ]

    for ri, row_data in enumerate(sample_scenes, 3):
        ws.row_dimensions[ri].height = 30
        for ci, val in enumerate(row_data, 1):
            if ci == 33:
                # Auto formula: Hook*0.45 + Than*0.35 + Ket*0.20
                cell = ws.cell(row=ri, column=ci)
                cell.value = f"=IFERROR(L{ri}*0.45+Y{ri}*0.35+AF{ri}*0.20,\"\")"
                cell.font = Font(size=8, name="Calibri", italic=True, color="1A237E", bold=True)
            else:
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(size=8, name="Calibri")
            cell.border = border_thin()
            cell.alignment = left_wrap() if ci in [3, 5, 10, 13, 16, 17, 19, 21, 26, 27, 30] else center()
            row_bg = C_WHITE if ri % 2 == 1 else C_LIGHT_GRAY
            cell.fill = fill(row_bg)

    # Blank rows for user
    for ri in range(4, 53):
        ws.row_dimensions[ri].height = 25
        row_bg = C_WHITE if ri % 2 == 0 else C_LIGHT_GRAY
        for ci in range(1, 34):
            cell = ws.cell(row=ri, column=ci)
            cell.fill = fill(row_bg)
            cell.border = border_thin()
            if ci == 33:
                cell.value = f"=IFERROR(L{ri}*0.45+Y{ri}*0.35+AF{ri}*0.20,\"\")"
                cell.font = Font(size=8, name="Calibri", italic=True, color="1A237E")
                cell.alignment = center()

    # Data validations
    for dv_range, opts in [
        ("K3:K52", '"Y,N"'),
        ("S3:S52", '"Y,N"'),
        ("U3:U52", '"Y,N"'),
        ("AC3:AC52", '"Y,N"'),
        ("AD3:AD52", '"Comment,Link bio,Swipe up,Tag bạn bè,DM,Mua ngay,Khác"'),
    ]:
        dv = DataValidation(type="list", formula1=opts, showDropDown=False)
        dv.sqref = dv_range
        ws.add_data_validation(dv)

    # Scoring number validations (1-10)
    dv_score = DataValidation(
        type="whole", operator="between", formula1="1", formula2="10"
    )
    dv_score.sqref = "L3:L52 Y3:Y52 AE3:AE52 AF3:AF52"
    ws.add_data_validation(dv_score)

    # Color scale for Hook scores
    ws.conditional_formatting.add(
        "L3:L52",
        ColorScaleRule(
            start_type="num", start_value=1, start_color="FF5252",
            mid_type="num",   mid_value=5,   mid_color="FFEB3B",
            end_type="num",   end_value=10,  end_color="4CAF50"
        )
    )

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: BẢNG ĐIỂM TỔNG HỢP
# ══════════════════════════════════════════════════════════════════════════════
def create_summary_sheet(wb):
    ws = wb.create_sheet("⭐ Bảng Điểm Tổng hợp")
    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False

    # Title
    merge_header(ws, 1, 1, 12, "⭐  BẢNG ĐIỂM TỔNG HỢP – VIDEO ANALYSIS ĐA NỀN TẢNG", C_HEADER_DARK, font_size=13)
    ws.row_dimensions[1].height = 30

    col_defs = [
        (1,  "STT",                  8,  C_HEADER_DARK),
        (2,  "Link Video",           35, C_HEADER_DARK),
        (3,  "Sản phẩm",             22, C_HEADER_MID),
        (4,  "Nền tảng",             13, C_HEADER_MID),
        (5,  "Ngày đăng",            13, C_HEADER_MID),
        (6,  "ER (%)",               12, "1565C0"),
        (7,  "Completion Rate (%)",  16, "1565C0"),
        (8,  "CTR Giỏ hàng (%)",     16, "1565C0"),
        (9,  "Điểm Hook",            13, "E53935"),
        (10, "Điểm Thân bài",        13, "1565C0"),
        (11, "Điểm Kết/CTA",         13, "2E7D32"),
        (12, "Điểm Nội dung",        14, C_ACCENT),
        (13, "Điểm Tổng Composite",  18, C_HEADER_DARK),
        (14, "Xếp hạng",             12, C_HEADER_DARK),
        (15, "Phân loại",            16, C_HEADER_MID),
    ]

    for col_idx, header, width, bg_color in col_defs:
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        cell.fill = fill(bg_color)
        cell.alignment = center()
        cell.border = border_thin()
    ws.row_dimensions[2].height = 35

    # Sample formulas referencing other sheets
    # For demo – 3 rows with references
    # Composite = ER*0.35 + CompRate*0.15 + CTR*0.15 + ContentScore*0.25 + Trend*0.10
    # Normalized to 0-10 scale (ER max ~10, CompRate max 100, CTR max ~5)
    notes = [
        "Công thức Điểm Tổng (Composite) = ER×0.35 + Completion%×0.15 + CTR×0.15 + DiemNoiDung×0.25 + DiemTrend×0.10",
        "Khung điểm giữ nguyên cho TikTok; với Facebook hãy diễn giải theo benchmark tham khảo ở sheet '📐 Thang điểm & Trọng số'",
        "Điểm Nội dung = Hook×0.45 + ThânBài×0.35 + Kết×0.20  (thang 1–10)",
    ]
    for ri, note in enumerate(notes, 3):
        ws.row_dimensions[ri].height = 22
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=15)
        cell = ws.cell(row=ri, column=1, value=note)
        cell.font = Font(size=9, name="Calibri", italic=True, color="1A237E")
        cell.fill = fill(C_SECTION_2)
        cell.border = border_thin()
        cell.alignment = left_wrap()

    # Data rows (blank, formula-driven – user will populate from other sheets)
    for ri in range(6, 53):
        ws.row_dimensions[ri].height = 20
        row_bg = C_WHITE if ri % 2 == 0 else C_LIGHT_GRAY
        for ci in range(1, 16):
            cell = ws.cell(row=ri, column=ci)
            cell.fill = fill(row_bg)
            cell.border = border_thin()
            cell.alignment = center()

    # Conditional formatting on composite score (col 13)
    ws.conditional_formatting.add(
        "M6:M52",
        ColorScaleRule(
            start_type="num", start_value=0,  start_color="FF5252",
            mid_type="num",   mid_value=5,    mid_color="FFEB3B",
            end_type="num",   end_value=10,   end_color="4CAF50"
        )
    )

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: THANG ĐIỂM & TRỌNG SỐ
# ══════════════════════════════════════════════════════════════════════════════
def create_scoring_sheet(wb):
    ws = wb.create_sheet("📐 Thang điểm & Trọng số")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 40

    merge_header(ws, 1, 2, 11, "📐  THANG ĐIỂM CHI TIẾT & TRỌNG SỐ CÁC TIÊU CHÍ", C_HEADER_DARK, font_size=13)
    ws.row_dimensions[1].height = 30

    def section(row, title, bg):
        ws.merge_cells(f"B{row}:F{row}")
        c = ws[f"B{row}"]
        c.value = title
        c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        c.fill = fill(bg)
        c.alignment = left_wrap()
        ws.row_dimensions[row].height = 22

    def header_row(row, cols_labels, bg):
        for ci, label in enumerate(cols_labels, 2):
            cell = ws.cell(row=row, column=ci, value=label)
            cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            cell.fill = fill(bg)
            cell.alignment = center()
            cell.border = border_thin()
        ws.row_dimensions[row].height = 30

    def data_row(row, values, bg=C_WHITE):
        ws.row_dimensions[row].height = 35
        for ci, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = left_wrap()

    def benchmark_section(row, title, bg):
        ws.merge_cells(f"H{row}:K{row}")
        c = ws[f"H{row}"]
        c.value = title
        c.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        c.fill = fill(bg)
        c.alignment = left_wrap()
        c.border = border_thin()
        for col in range(8, 12):
            ws.cell(row=row, column=col).border = border_thin()
        ws.row_dimensions[row].height = 22

    def benchmark_header(row, labels, bg):
        for ci, label in enumerate(labels, 8):
            cell = ws.cell(row=row, column=ci, value=label)
            cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            cell.fill = fill(bg)
            cell.alignment = center()
            cell.border = border_thin()
        ws.row_dimensions[row].height = 32

    def benchmark_row(row, values, bg=C_WHITE):
        ws.row_dimensions[row].height = 42
        for ci, val in enumerate(values, 8):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = left_wrap()

    # ── ENGAGEMENT METRICS ──
    section(3, "C. CHỈ SỐ TƯƠNG TÁC (Trọng số nhóm: 35%)", "1565C0")
    header_row(4, ["Tiêu chí", "Trọng số", "Ngưỡng điểm", "Mô tả chi tiết", "Benchmark ngành"], "1565C0")
    eng_data = [
        ["Lượt xem (Views)", "8%",
         "1=<1K | 3=1K-10K | 5=10K-50K | 7=50K-200K | 9=200K-1M | 10=>1M",
         "Chỉ số reach cơ bản. Video viral thường >500K views trong 48h đầu.",
         "Tốt: >50K | Rất tốt: >200K"],
        ["Engagement Rate – ER", "10%",
         "1=<0.5% | 3=0.5-1% | 5=1-3% | 7=3-5% | 9=5-10% | 10=>10%",
         "ER = (Likes+Comments+Shares+Saves)/Views×100%. Đây là chỉ số quan trọng nhất phản ánh chất lượng nội dung.",
         "Tốt: >3% | Rất tốt: >5%"],
        ["Completion Rate", "8%",
         "1=<10% | 3=10-20% | 5=20-35% | 7=35-50% | 9=50-70% | 10=>70%",
         "Tỉ lệ người xem hết video. Completion rate cao = nội dung hấp dẫn từ đầu đến cuối.",
         "Tốt: >30% | Rất tốt: >50%"],
        ["CTR Giỏ hàng", "9%",
         "1=<0.3% | 3=0.3-0.7% | 5=0.7-1.5% | 7=1.5-3% | 9=3-5% | 10=>5%",
         "Tỉ lệ nhấp link mua hàng. Chỉ số này phản ánh trực tiếp hiệu quả bán hàng của video.",
         "Tốt: >1% | Rất tốt: >3%"],
        ["Lượt lưu (Saves)", "5%",
         "1=<10 | 3=10-50 | 5=50-200 | 7=200-1K | 9=1K-5K | 10=>5K",
         "Lượt lưu cho thấy nội dung có giá trị tham khảo lâu dài – đặc biệt quan trọng cho sản phẩm dinh dưỡng.",
         "Tốt: >500 | Rất tốt: >2K"],
    ]
    for ri, row_vals in enumerate(eng_data, 5):
        data_row(ri, row_vals, C_WHITE if ri % 2 == 1 else C_SECTION_2)

    benchmark_section(3, "📘 BENCHMARK TIKTOK (khung hiện có)", "1565C0")
    benchmark_header(4, ["Chỉ số", "Mức tốt", "Mức rất tốt", "Ghi chú"], "1565C0")
    tiktok_benchmark_rows = [
        ["ER", ">3%", ">5%", "Phù hợp video organic/shop content ngắn, nhịp nhanh."],
        ["Completion Rate", ">30%", ">50%", "Video ngắn có hook mạnh thường giữ chân tốt hơn."],
        ["CTR CTA/Giỏ hàng", ">1%", ">3%", "Áp dụng khi có link bio, TikTok Shop hoặc CTA mua hàng rõ."],
        ["Saves", ">500", ">2K", "Đặc biệt hữu ích với nội dung tips, công thức, checklist."],
    ]
    for ri, row_vals in enumerate(tiktok_benchmark_rows, 5):
        benchmark_row(ri, row_vals, C_WHITE if ri % 2 == 1 else C_SECTION_2)

    # ── GROWTH METRICS ──
    section(11, "D. CHỈ SỐ TĂNG TRƯỞNG (Trọng số nhóm: 10%)", "6A1B9A")
    header_row(12, ["Tiêu chí", "Trọng số", "Ngưỡng điểm", "Mô tả chi tiết", "Benchmark ngành"], "6A1B9A")
    growth_data = [
        ["Follow mới từ video", "5%",
         "1=<5 | 3=5-20 | 5=20-100 | 7=100-500 | 9=500-2K | 10=>2K",
         "Số lượng follow mới nhờ một video. Quan trọng giai đoạn xây kênh từ 0.",
         "Tốt: >100 | Rất tốt: >500"],
        ["Điểm bắt trend (1-10)", "5%",
         "1=Không trend | 4=Trend cũ | 7=Đang trend | 10=Là người tạo trend",
         "Đánh giá chủ quan mức độ video khai thác xu hướng âm nhạc/challenge/chủ đề viral.",
         "Nên bắt trend khi độ phủ >70% TikTok VN"],
    ]
    for ri, row_vals in enumerate(growth_data, 13):
        data_row(ri, row_vals, C_WHITE if ri % 2 == 1 else C_SECTION_5)

    benchmark_section(11, "📗 BENCHMARK FACEBOOK (ước lượng tham khảo theo thông lệ ngành)", "6A1B9A")
    benchmark_header(12, ["Chỉ số", "Mức tốt", "Mức rất tốt", "Ghi chú"], "6A1B9A")
    facebook_benchmark_rows = [
        ["ER", "≈1-2%", ">2.5-3%", "Facebook thường có ER thấp hơn TikTok với cùng ngành; nên so trong cùng placement."],
        ["Completion / Watch Rate", "≈15-25%", ">25-35%", "Video Feed/Watch dài hơn nên tỉ lệ xem hết thấp hơn; có thể tham chiếu ThruPlay hoặc watch rate."],
        ["CTR CTA / Link", "≈0.5-1.0%", ">1.0-1.8%", "Dùng link CTR, outbound CTR hoặc CTA click rate tùy báo cáo Meta."],
        ["3-second video views / Retention sớm", "Theo dõi xu hướng tăng", "Tăng rõ theo test", "Chỉ số chẩn đoán sớm cho hook; không thay thế hoàn toàn completion rate."],
    ]
    for ri, row_vals in enumerate(facebook_benchmark_rows, 13):
        benchmark_row(ri, row_vals, C_WHITE if ri % 2 == 1 else C_SECTION_5)

    # ── HOOK ──
    section(16, "E. CHẤT LƯỢNG NỘI DUNG – HOOK (Trọng số nhóm: 25% | trong Điểm Nội dung: 45%)", "E53935")
    header_row(17, ["Tiêu chí", "Trọng số", "Ngưỡng điểm", "Mô tả chi tiết", "Ví dụ tốt"], "E53935")
    hook_data = [
        ["Điểm Hook tổng (1-10)", "25%",
         "1-3=Nhàm chán | 4-5=Bình thường | 6-7=Tốt | 8-9=Rất tốt | 10=Xuất sắc",
         "Đánh giá tổng thể 5s đầu: có gây tò mò không? Có đặt câu hỏi/shock/pain point không? Với Facebook có thể đánh giá linh hoạt trong 3–5 giây đầu do nhịp lướt chậm hơn.",
         '"Ăn chay 3 tháng, da tôi thay đổi không ngờ..." / Cầm đồ ăn vứt vào thùng gây shock'],
        ["Pain point rõ ràng", "—",
         "Y = có nêu pain point rõ ràng ngay 5s đầu\nN = không có",
         "Nêu vấn đề khách hàng đang gặp phải: thiếu năng lượng, da xấu, thiếu protein khi chay,...",
         '"Bạn ăn chay mà luôn mệt mỏi? Đây là lý do..."'],
        ["Lời thoại gây tò mò", "—",
         "Câu hỏi mở / câu khẳng định gây shock / con số bất ngờ",
         "Script hook quan trọng: phải có móc câu rõ ràng trong 3 giây đầu tiên.",
         '"90% người ăn chay đang thiếu chất này mà không biết"'],
        ["Góc máy & Bối cảnh", "—",
         "Close-up mặt = tạo kết nối | Overhead = demo sản phẩm | Wide = lifestyle",
         "Góc máy close-up mặt người dùng tăng empathy. Bối cảnh sạch, ánh sáng tốt.",
         "Close-up mặt người dùng kết hợp sản phẩm cầm tay"],
    ]
    for ri, row_vals in enumerate(hook_data, 18):
        data_row(ri, row_vals, C_WHITE if ri % 2 == 0 else C_SECTION_4)

    # ── THÂN BÀI ──
    section(23, "F. CHẤT LƯỢNG NỘI DUNG – THÂN BÀI (Trọng số nhóm: 20% | trong Điểm ND: 35%)", "1565C0")
    header_row(24, ["Tiêu chí", "Trọng số", "Ngưỡng điểm", "Mô tả chi tiết", "Ví dụ tốt"], "1565C0")
    body_data = [
        ["Điểm Thân bài (1-10)", "20%",
         "1-3=Rối rắm | 4-5=Thiếu logic | 6-7=Rõ ràng | 8-9=Thuyết phục | 10=Chuyển đổi cao",
         "Đánh giá mạch nội dung: có logic không, có bằng chứng không, có giữ được sự chú ý không.",
         "Giải thích 3 lợi ích + dẫn review khách hàng + demo trực tiếp"],
        ["Bằng chứng / Social proof", "—",
         "0=Không có | 1=Tuyên bố suông | 2=Review | 3=Số liệu+Review+Demo",
         "Social proof giúp tăng niềm tin: review video, số lượng người dùng, chứng chỉ,...",
         '"Hơn 10,000 khách hàng đã tin dùng – review thật 100%"'],
        ["So sánh sản phẩm", "—",
         "Y = có so sánh với đối thủ hoặc sản phẩm khác\nN = không",
         "So sánh giúp định vị sản phẩm: so với whey protein thường, collagen từ động vật,...",
         "Bảng so sánh: Vegan Collagen vs Collagen cá – độ hấp thụ, nguồn gốc, đạo đức"],
        ["Điểm gãy (nói dông dài)", "—",
         "Y = có đoạn dông dài làm người xem rời đi\nN = không",
         "Xác định timestamp mà Completion Rate giảm mạnh = điểm gãy cần cắt.",
         "Nếu >15% người xem rời đi tại giây X → cắt/edit lại đoạn đó"],
    ]
    for ri, row_vals in enumerate(body_data, 25):
        data_row(ri, row_vals, C_WHITE if ri % 2 == 1 else C_SECTION_2)

    # ── KẾT / CTA ──
    section(30, "G. CHẤT LƯỢNG NỘI DUNG – KẾT / CTA (Trọng số nhóm: 10% | trong Điểm ND: 20%)", "2E7D32")
    header_row(31, ["Tiêu chí", "Trọng số", "Ngưỡng điểm", "Mô tả chi tiết", "Ví dụ tốt"], "2E7D32")
    cta_data = [
        ["Điểm Kết/CTA (1-10)", "10%",
         "1-3=Không có CTA | 4-5=CTA mờ | 6-7=CTA rõ | 8-9=CTA + urgency | 10=CTA + urgency + offer",
         "CTA phải rõ ràng, có urgency và kèm offer nếu có. Không để người xem đoán phải làm gì.",
         '"Mua ngay hôm nay – giảm 30% chỉ còn 24h – link bio"'],
        ["Loại CTA", "—",
         "Phân loại: Link bio | Comment từ khóa | DM | Swipe up | Tag bạn | Follow",
         "Mỗi loại CTA phù hợp với mục tiêu khác nhau. Bán hàng nên dùng Link bio hoặc Comment từ khóa.",
         'Comment "COLLAGEN" để nhận báo giá – tăng engagement và warm lead'],
        ["Urgency / Scarcity", "—",
         "Y = có yếu tố khan hiếm / giới hạn thời gian\nN = không",
         "Urgency tăng tỉ lệ click: 'chỉ còn 50 suất', 'flash sale 2h', 'ưu đãi hết ngày hôm nay'.",
         '"Flash sale – chỉ còn 2 tiếng – đừng bỏ lỡ"'],
    ]
    for ri, row_vals in enumerate(cta_data, 32):
        data_row(ri, row_vals, C_WHITE if ri % 2 == 0 else C_SECTION_1)

    # ── CÔNG THỨC TƯƠNG QUAN ──
    section(36, "📊  CÔNG THỨC & PHÂN TÍCH TƯƠNG QUAN", C_HEADER_DARK)
    header_row(37, ["Phân tích", "Công thức Excel", "Diễn giải", "Ứng dụng", "Ngưỡng ý nghĩa"], C_HEADER_DARK)
    corr_data = [
        ["Tương quan Hook vs CTR",
         "=CORREL('📋 Dữ liệu Video'!Q3:Q52,'🎬 Phân tích Phân cảnh'!L3:L52)",
         "r > 0.6 = Hook ảnh hưởng mạnh đến CTR. r < 0.3 = không tương quan.",
         "Nếu tương quan cao → ưu tiên đầu tư vào Hook hơn nội dung thân bài.",
         "r > 0.6 = tương quan cao"],
        ["Tương quan Completion Rate vs ER",
         "=CORREL('📋 Dữ liệu Video'!R3:R52,'📋 Dữ liệu Video'!S3:S52)",
         "Video được xem hết thường có ER cao hơn. Completion rate ảnh hưởng đến thuật toán.",
         "Nếu r > 0.7 → tập trung cải thiện khả năng giữ chân người xem.",
         "r > 0.5 = đáng chú ý"],
        ["Video viral tốt nhất",
         "=LARGE('⭐ Bảng Điểm Tổng hợp'!M6:M52,1)",
         "Tìm video có Điểm Tổng cao nhất để nhân bản công thức.",
         "Phân tích video top để tìm pattern: thể loại, thời lượng, sản phẩm...",
         "Top 20% video tạo ra 80% doanh thu"],
        ["Điểm trung bình theo sản phẩm",
         "=AVERAGEIF('📋 Dữ liệu Video'!H3:H52,\"Tên SP\",'⭐ Bảng Điểm Tổng hợp'!M6:M52)",
         "So sánh hiệu quả video theo từng sản phẩm. Sản phẩm nào có điểm cao nhất?",
         "Ưu tiên sản xuất content cho sản phẩm có điểm trung bình cao nhất.",
         "Chênh >1.5 điểm = đáng kể"],
        ["Xu hướng theo mùa",
         "=AVERAGEIFS(DiemTong,NgayDang,\">=2026-01-01\",NgayDang,\"<=2026-03-31\")",
         "So sánh điểm trung bình theo quý để tìm mùa bán hàng tốt nhất.",
         "Lên kế hoạch sản xuất content theo mùa cao điểm đã xác định.",
         "Chênh >15% là xu hướng rõ"],
        ["So sánh TikTok vs Facebook",
         "=AVERAGEIF('📋 Dữ liệu Video'!D3:D52,\"TikTok\",'⭐ Bảng Điểm Tổng hợp'!M6:M52)-AVERAGEIF('📋 Dữ liệu Video'!D3:D52,\"Facebook Reels\",'⭐ Bảng Điểm Tổng hợp'!M6:M52)",
         "Đo chênh lệch composite giữa 2 nền tảng trên cùng khung chấm điểm.",
         "Dùng để xem concept nào nên ưu tiên nhân bản sang Facebook hoặc giữ riêng cho TikTok.",
         "Nên so trong cùng sản phẩm/cùng loại quảng cáo"],
    ]
    for ri, row_vals in enumerate(corr_data, 38):
        data_row(ri, row_vals, C_WHITE if ri % 2 == 0 else C_SECTION_2)

    benchmark_section(18, "📝 QUY ĐỔI CHỈ SỐ META → KHUNG NÀY", C_HEADER_DARK)
    benchmark_header(19, ["Nguồn Meta", "Nhập vào cột", "Mức ưu tiên", "Ghi chú"], C_HEADER_DARK)
    mapping_rows = [
        ["ThruPlay / video plays at 95%+", "Completion Rate (%)", "Ưu tiên cao", "Dùng khi campaign tối ưu view/video completion."],
        ["Average Watch Time + tỉ lệ xem các mốc", "Completion Rate (%)", "Ưu tiên trung bình", "Nếu không có completion trực tiếp, dùng để ước lượng chất lượng giữ chân."],
        ["Link CTR / Outbound CTR", "CTR Giỏ hàng (%)", "Ưu tiên cao", "Chọn đúng chỉ số gần nhất với hành vi click CTA."],
        ["Love/Haha/Wow/Sad/Angry", "Ghi chú định tính", "Bổ trợ", "Không cộng riêng vào composite nhưng hữu ích để đọc cảm xúc phản hồi."],
    ]
    for ri, row_vals in enumerate(mapping_rows, 20):
        benchmark_row(ri, row_vals, C_WHITE if ri % 2 == 0 else C_SECTION_2)

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    create_guide_sheet(wb)
    create_data_sheet(wb)
    create_scene_sheet(wb)
    create_summary_sheet(wb)
    create_scoring_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"✅ Excel file saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
